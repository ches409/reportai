import os
import json
import asyncio
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Tuple, Union
from dataclasses import dataclass, asdict, field
from pathlib import Path
import logging
import re
import copy
from collections import Counter

# 라이브러리
from notion_client import AsyncClient as NotionClient
import httpx  # Ollama API 호출용
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from weasyprint import HTML #pdf생성 라이브러리
import zipfile
from cryptography.fernet import Fernet
import boto3
import requests
from fastapi import FastAPI, BackgroundTasks
from pydantic import BaseModel
import uvicorn
import logging
from dotenv import load_dotenv

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

# 환경변수 로드
load_dotenv()

# 로거 설정
logger = logging.getLogger("uvicorn")
logger.setLevel(logging.DEBUG)

# 프롬프트 템플릿 로드
BASE_DIR = Path(__file__).resolve().parent
filter_select_prompt_path = BASE_DIR / "filter_select_prompt"
table_select_prompt_path = BASE_DIR / "table_select_prompt"

with filter_select_prompt_path.open("r", encoding="utf-8") as f:
    FILTER_SELECT_PROMPT = f.read()

with table_select_prompt_path.open("r", encoding="utf-8") as f:
    TABLE_SELECT_PROMPT = f.read()



class Config:
    NOTION_TOKEN: str = os.getenv("NOTION_TOKEN", "")

    DB_CLASS: str = os.getenv("DB_CLASS", "")
    DB_REPORTREQUEST: str = os.getenv("DB_REPORTREQUEST", "")
    DB_DISCHARGE: str = os.getenv("DB_DISCHARGE", "")
    DB_STUDENT: str = os.getenv("DB_STUDENT", "")

    OLLAMA_URL: str = os.getenv("OLLAMA_URL", "http://localhost:11434")
    OLLAMA_ENTITY_MODEL: str = os.getenv("OLLAMA_ENTITY_MODEL", "qwen3:8b")
    OLLAMA_QUERY_MODEL: str = os.getenv("OLLAMA_QUERY_MODEL", "qwencoder:7b")

    TEMP_DIR = Path("temp")
    REPORTS_DIR = Path("reports")
    #ENCRYPTION_KEY = os.getenv("ENCRYPTION_KEY", Fernet.generate_key())

    def __post_init__(self):
        self.TEMP_DIR.mkdir(exist_ok=True)
        self.REPORTS_DIR.mkdir(exist_ok=True)

config = Config()

#### 데이터 클래스 정의

@dataclass
class Class:
    id: str
    student_name: str
    teacher_name: list[str]
    class_name: str
    parent_phone_number: str
    start_date: datetime
    school_name: str
    grade: int
    type: str  # 특목, 본관 구분

@dataclass
class DISCHARGE:
    id: str
    student_name: str
    teacher_name: list[str]
    class_name: str
    parent_phone_number: str
    student_phone_number: str
    discharge_date: datetime
    start_date: datetime
    discharging_reason: str
    school_name: str
    grade: int
    type: str  # 특목, 본관 구분



@dataclass
class ReportRequest:
    id: str
    question: str
    requester_name: str
    status: str  # 대기중, 처리중, 완료, 실패
    created_at: datetime
    updated_at: datetime

@dataclass
class ReportQuery:
    target_table: Optional[str] = None 
    filters: Dict[str, any] = field(default_factory=dict)
    columns: List[str] = field(default_factory=list)
    aggregations: Optional[List[str]] = None
    sort_by: Optional[str] = None
    date_range: Optional[Dict[str, str]] = None


####


class NotionManager:
    def __init__(self):
        self.client = NotionClient(auth=config.NOTION_TOKEN)
        self.db_map = {
            "class": config.DB_CLASS,
            "report_requests": config.DB_REPORTREQUEST,
            "discharge": config.DB_DISCHARGE,
            "student": config.DB_STUDENT
        }

    async def get_pending_requests(self) -> List[ReportRequest]:
        logger.info("📋 보고서 요청 DB 확인 중...")
    
        response = await self.client.databases.query(
            database_id=self.db_map["report_requests"],
            filter={
                "property": "상태",
                "status": {"equals": "대기중"}
            }
        )

        requests = []
        for page in response["results"]:
            req = ReportRequest(
                id=page["id"],
                question=self._get_title(page, "질문"),
                requester_name=self._get_person_name(page, "요청자"),
                status="대기중",
                created_at=datetime.fromisoformat(page["created_time"].replace("Z", "+00:00")),
                updated_at=datetime.fromisoformat(page["last_edited_time"].replace("Z", "+00:00"))
            )
            requests.append(req)
        
        if requests:
            logger.info(f"✅ {len(requests)}개의 요청 발견")
        return requests
    
    async def query_table(self, table_name: str, query: ReportQuery) -> List[Dict]:
        logger.info(f"📊 {table_name} 테이블 조회 중...")
        
        # allow case-insensitive table name lookup (AI may return 'Class' or 'DISCHARGE')
        db_id = self.db_map.get(table_name)
        if not db_id:
            db_id = self.db_map.get(table_name.lower())
        if not db_id:
            lower_map = {k.lower(): v for k, v in self.db_map.items()}
            db_id = lower_map.get(table_name.lower())
        if not db_id:
            logger.error(f"테이블을 찾을 수 없음: {table_name}")
            return []
        logger.debug(f"query_table: resolved '{table_name}' -> db_id '{db_id}'")
        
        # 필터 생성
        notion_filter = self._build_filter(query)
        
        # 페이지네이션 처리: Notion이 결과를 여러 페이지로 반환할 수 있음
        all_results = []
        start_cursor = None
        while True:
            response = await self.client.databases.query(
                database_id=db_id,
                filter=notion_filter if notion_filter else None,
                start_cursor=start_cursor
            )
            results = response.get("results", [])
            all_results.extend(results)

            if not response.get("has_more"):
                break
            start_cursor = response.get("next_cursor")

        # 데이터 파싱
        data = []
        for page in all_results:
            row = {}
            for col in query.columns:
                row[col] = self._extract_property(page, col)
            data.append(row)
        
        # 날짜 필터 추가 적용 (Notion API 필터가 완벽하지 않을 수 있으므로)
        if query.date_range and isinstance(query.date_range, dict) and query.date_range.get("property"):
            date_prop = query.date_range.get("property")
            start_val = query.date_range.get("start")
            end_val = query.date_range.get("end")
            
            if start_val or end_val:
                filtered_data = []
                for row in data:
                    date_str = row.get(date_prop, "")
                    if not date_str:
                        continue
                    
                    try:
                        # 날짜 문자열 파싱 (ISO 형식 또는 YYYY-MM-DD)
                        if "T" in date_str:
                            row_date = datetime.fromisoformat(date_str.split("T")[0])
                        else:
                            row_date = datetime.fromisoformat(date_str)
                        
                        # 날짜 범위 체크
                        if start_val:
                            start_date = datetime.fromisoformat(start_val)
                            if row_date < start_date:
                                continue
                        
                        if end_val:
                            end_date = datetime.fromisoformat(end_val)
                            if row_date > end_date:
                                continue
                        
                        filtered_data.append(row)
                    except Exception as e:
                        logger.warning(f"날짜 파싱 실패 (행 건너뜀): {date_str}, 오류: {str(e)}")
                        continue
                
                logger.info(f"📅 날짜 필터 적용: {len(data)}건 → {len(filtered_data)}건 (범위: {start_val} ~ {end_val})")
                data = filtered_data

        logger.info(f"✅ {len(data)}건 조회 완료 (pages: {len(all_results)})")
        return data
    

    async def query_multiple_tables(self, queries: List[ReportQuery]) -> Dict[str, List[Dict]]:
        # allow single ReportQuery or list of them
        if isinstance(queries, ReportQuery):
            queries = [queries]

        table_names = []
        for q in queries:
            try:
                table_names.append(str(q.target_table))
            except Exception:
                table_names.append("<unknown>")

        logger.info(f"🔗 멀티 테이블 조회: {', '.join(table_names)}")

        # 각 테이블별로 데이터 조회
        all_data = {}
        for query in queries:
            table_name = query.target_table
            # clone query per-table so we can inject a sensible default date property
            q_clone = copy.deepcopy(query)

            # If AI provided a date_range but omitted the property name, set per-table defaults
            if q_clone.date_range and isinstance(q_clone.date_range, dict) and not q_clone.date_range.get("property"):
                default_prop = None
                if str(table_name).lower() == "class":
                    default_prop = "start_date"
                elif str(table_name).lower() == "discharge":
                    default_prop = "discharge_date"

                if default_prop:
                    q_clone.date_range["property"] = default_prop
                    logger.debug(f"query_multiple_tables: set default date property '{default_prop}' for table '{table_name}'")

            data = await self.query_table(table_name, q_clone)
            all_data[table_name] = data

        # If only one table requested, return its data under its table name
        return all_data
    
    def _join_tables(self, all_data: Dict[str, List[Dict]], join_key: str) -> List[Dict]:
        logger.info(f"🔗 조인 키: {join_key}")
        
        # 첫 번째 테이블을 기준으로
        base_table = list(all_data.keys())[0]
        result = []
        
        for base_row in all_data[base_table]:
            joined_row = base_row.copy()
            join_value = base_row.get(join_key)
            
            if not join_value:
                result.append(joined_row)
                continue
            
            # 다른 테이블에서 매칭되는 데이터 찾기
            for table_name, table_data in all_data.items():
                if table_name == base_table:
                    continue
                
                # 해당 테이블에서 조인 키 값이 일치하는 행 찾기
                matching_rows = [
                    row for row in table_data 
                    if row.get(join_key) == join_value
                ]
                
                # 매칭된 데이터 병합 (컬럼명 충돌 방지)
                for match in matching_rows:
                    for key, value in match.items():
                        if key != join_key:  # 조인 키는 중복 제거
                            new_key = f"{table_name}_{key}"
                            joined_row[new_key] = value
            
            result.append(joined_row)
        
        logger.info(f"✅ 조인 완료: {len(result)}건")
        return result



    def _build_filter(self, query: ReportQuery) -> Optional[Dict]:
        conditions = []
        
        # 일반 필터 처리
        if query.filters:
            for key, value in query.filters.items():
                if isinstance(value, str):
                    conditions.append({
                        "property": key,
                        "rich_text": {"contains": value}
                    })
                elif isinstance(value, (int, float)):
                    conditions.append({
                        "property": key,
                        "number": {"equals": value}
                    })
                elif isinstance(value, list):
                    conditions.append({
                        "property": key,
                        "select": {"equals": value[0]}
                    })
        
        # 날짜 필터는 query_table에서 처리 (Notion API 필터가 완벽하지 않을 수 있으므로)
        
        # 조건이 하나도 없으면 None 반환
        if len(conditions) == 0:
            return None
        elif len(conditions) > 1:
            return {"and": conditions}
        else:
            return conditions[0]
    
    async def update_request_status(self, request_id: str, status: str,
                                    error: str = None):
        properties = {
            "상태": {"status": {"name": status}},
            "완료일": {"date": {"start": datetime.now().isoformat()}}
        }        
        
        if error:
            properties["비고"] = {
                "rich_text": [{"text": {"content": f"에러: {error}"}}]
            }
        
        await self.client.pages.update(page_id=request_id, properties=properties)
        logger.info(f"✅ 상태 업데이트: {status}")
    
    # 헬퍼 메서드
    def _get_title(self, page: Dict, prop: str) -> str:
        p = page["properties"].get(prop, {})
        return p["title"][0]["text"]["content"] if p.get("title") else ""
    
    def _get_person_name(self, page: Dict, prop: str) -> str:
        p = page["properties"].get(prop, {})
        return p["people"][0]["name"] if p.get("people") else ""
    
    def _get_person_email(self, page: Dict, prop: str) -> str:
        p = page["properties"].get(prop, {})
        return p["people"][0].get("person", {}).get("email", "") if p.get("people") else ""
    
    def _extract_property(self, page: Dict, prop: str):
        p = page["properties"].get(prop, {})
        prop_type = p.get("type", "")
        
        if prop_type == "title":
            return p["title"][0]["text"]["content"] if p.get("title") else ""
        elif prop_type == "rich_text":
            return p["rich_text"][0]["text"]["content"] if p.get("rich_text") else ""
        elif prop_type == "number":
            return p.get("number", 0)
        elif prop_type == "select":
            return p["select"]["name"] if p.get("select") else ""
        elif prop_type == "date":
            return p["date"]["start"] if p.get("date") else ""
        elif prop_type == "phone_number":
            return p.get("phone_number", "")
        return ""
    
    def _get_select(self, page: Dict, prop: str):
        p = page["properties"].get(prop, {})
        return p["select"]["name"] if p.get("select") else ""

    def _get_multi_select(self, page: Dict, prop: str):
        p = page["properties"].get(prop, {})
        return [v["name"] for v in p.get("multi_select", [])]

    def _get_number(self, page: Dict, prop: str):
        return page["properties"].get(prop, {}).get("number")

    def _get_date(self, page: Dict, prop: str):
        p = page["properties"].get(prop, {})
        return p["date"]["start"] if p.get("date") else None
    
    def _get_rich_text_value(self, page: Dict, prop: str) -> str:
        p = page["properties"].get(prop, {})
        if p.get("rich_text"):
            return p["rich_text"][0]["plain_text"].strip()
        return ""
    
    async def get_date_range_from_table(self, table_name: str, date_property: Optional[str] = None) -> Optional[Tuple[datetime, datetime]]:
        """테이블에서 날짜 범위 조회 (첫 날짜와 마지막 날짜)
        
        Args:
            table_name: 테이블 이름 (class, discharge, student)
            date_property: 날짜 속성명 (None이면 자동 결정)
                - class, student: start_date
                - discharge: discharge_date
        
        Returns:
            (첫 날짜, 마지막 날짜) 튜플 또는 None (데이터가 없을 경우)
        """
        db_id = self.db_map.get(table_name.lower())
        if not db_id:
            logger.error(f"❌ 테이블을 찾을 수 없음: {table_name}")
            return None
        
        # 날짜 속성 자동 결정
        if not date_property:
            if table_name.lower() in ["class", "student"]:
                date_property = "start_date"
            elif table_name.lower() == "discharge":
                date_property = "discharge_date"
            else:
                logger.error(f"❌ 알 수 없는 테이블 타입: {table_name}")
                return None
        
        try:
            # 첫 번째 날짜 조회 (오름차순)
            first_result = await self.client.databases.query(
                database_id=db_id,
                sorts=[{"property": date_property, "direction": "ascending"}],
                page_size=1
            )
            
            # 마지막 날짜 조회 (내림차순)
            last_result = await self.client.databases.query(
                database_id=db_id,
                sorts=[{"property": date_property, "direction": "descending"}],
                page_size=1
            )
            
            first_date = None
            last_date = None
            
            if first_result.get('results') and len(first_result['results']) > 0:
                first_date_str = self._get_date(first_result['results'][0], date_property)
                if first_date_str:
                    try:
                        if "T" in first_date_str:
                            first_date = datetime.fromisoformat(first_date_str.split("T")[0])
                        else:
                            first_date = datetime.fromisoformat(first_date_str)
                    except Exception as e:
                        logger.error(f"❌ 첫 날짜 파싱 실패: {e}")
            
            if last_result.get('results') and len(last_result['results']) > 0:
                last_date_str = self._get_date(last_result['results'][0], date_property)
                if last_date_str:
                    try:
                        if "T" in last_date_str:
                            last_date = datetime.fromisoformat(last_date_str.split("T")[0])
                        else:
                            last_date = datetime.fromisoformat(last_date_str)
                    except Exception as e:
                        logger.error(f"❌ 마지막 날짜 파싱 실패: {e}")
            
            if first_date and last_date:
                logger.info(f"📅 [{table_name}] 날짜 범위: {first_date.date()} ~ {last_date.date()}")
                return (first_date, last_date)
            elif first_date:
                # 데이터가 하나만 있는 경우
                logger.info(f"📅 [{table_name}] 날짜: {first_date.date()}")
                return (first_date, first_date)
            else:
                logger.info(f"📭 [{table_name}] 테이블에 데이터가 없습니다.")
                return None
                
        except Exception as e:
            logger.error(f"❌ 날짜 범위 조회 실패: {e}")
            return None




####

class OllamaAnalyzer:
    def __init__(self):
        self.url = f"{config.OLLAMA_URL}/api/generate"
        self.model = config.OLLAMA_ENTITY_MODEL
    
    def _parse_date_range(self, question: str) -> Optional[Dict[str, str]]:
        """질문에서 날짜 범위를 파싱하여 반환"""
        question_lower = question.lower()
        now = datetime.now()
        
        # "X년 Y월부터 Z월까지" 형식 처리 (예: "2025년 3월부터 7월까지", "2025년 유형신 선생님 3월부터 7월까지")
        # 더 유연한 패턴: 년도와 첫 번째 월 사이, "부터"와 "까지" 사이에 어떤 텍스트가 있어도 매칭
        month_range_match = re.search(r'(\d{4})\s*년.*?(\d{1,2})\s*월\s*부터.*?(\d{1,2})\s*월\s*까지', question)
        if month_range_match:
            year = int(month_range_match.group(1))
            start_month = int(month_range_match.group(2))
            end_month = int(month_range_match.group(3))
            start_date = datetime(year, start_month, 1)
            # 종료 월의 마지막 날 계산
            if end_month == 12:
                end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(year, end_month + 1, 1) - timedelta(days=1)
            result = {
                "start": start_date.strftime("%Y-%m-%d"),
                "end": end_date.strftime("%Y-%m-%d")
            }
            logger.info(f"📅 날짜 범위 파싱 (월 범위): {year}년 {start_month}월 ~ {end_month}월 → {result['start']} ~ {result['end']}")
            return result
        
        # "X년 Y월부터" 형식 처리 (예: "2025년 3월부터", "2025년 유형신 선생님 3월부터")
        # 년도와 월 사이에 어떤 텍스트가 있어도 매칭
        month_start_match = re.search(r'(\d{4})\s*년.*?(\d{1,2})\s*월\s*부터', question)
        if month_start_match:
            year = int(month_start_match.group(1))
            start_month = int(month_start_match.group(2))
            start_date = datetime(year, start_month, 1)
            # 현재 날짜까지 또는 해당 년도 말까지
            end_date = datetime(year, 12, 31)
            if year == now.year and start_month <= now.month:
                end_date = now
            return {
                "start": start_date.strftime("%Y-%m-%d"),
                "end": end_date.strftime("%Y-%m-%d")
            }
        
        # "X년 Y월" 형식 처리 (예: "2025년 3월", "2025년 유형신 선생님 3월")
        # 년도와 월 사이에 어떤 텍스트가 있어도 매칭 (단, "부터"나 "까지"가 바로 뒤에 오는 경우는 제외)
        single_month_match = re.search(r'(\d{4})\s*년.*?(\d{1,2})\s*월\s*(?!부터|까지)', question)
        if single_month_match:
            year = int(single_month_match.group(1))
            month = int(single_month_match.group(2))
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(year, month + 1, 1) - timedelta(days=1)
            return {
                "start": start_date.strftime("%Y-%m-%d"),
                "end": end_date.strftime("%Y-%m-%d")
            }
        
        # 년도 추출 (예: "2024년", "2023년도") - 월 범위가 아닌 경우만
        year_match = re.search(r'(\d{4})\s*년(?!\s*\d)', question)
        if year_match:
            year = int(year_match.group(1))
            return {
                "start": f"{year}-01-01",
                "end": f"{year}-12-31"
            }
        
        # "올해", "이번 년", "올해 전체"
        if any(keyword in question_lower for keyword in ["올해", "이번 년", "올해 전체", "올해 전체"]):
            year = now.year
            return {
                "start": f"{year}-01-01",
                "end": f"{year}-12-31"
            }
        
        # "작년", "작년도"
        if any(keyword in question_lower for keyword in ["작년", "작년도"]):
            year = now.year - 1
            return {
                "start": f"{year}-01-01",
                "end": f"{year}-12-31"
            }
        
        # "이번 달", "이번월", "이번 달", "이번 월"
        if any(keyword in question_lower for keyword in ["이번 달", "이번월", "이번 달", "이번 월", "이번달"]):
            year = now.year
            month = now.month
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(year, month + 1, 1) - timedelta(days=1)
            return {
                "start": start_date.strftime("%Y-%m-%d"),
                "end": end_date.strftime("%Y-%m-%d")
            }
        
        # "지난 달", "지난달", "저번 달"
        if any(keyword in question_lower for keyword in ["지난 달", "지난달", "저번 달", "저번달"]):
            if now.month == 1:
                year = now.year - 1
                month = 12
            else:
                year = now.year
                month = now.month - 1
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(year, month + 1, 1) - timedelta(days=1)
            return {
                "start": start_date.strftime("%Y-%m-%d"),
                "end": end_date.strftime("%Y-%m-%d")
            }
        
        # "최근 N개월", "최근 N달"
        recent_match = re.search(r'최근\s*(\d+)\s*개?월', question_lower)
        if recent_match:
            months = int(recent_match.group(1))
            end_date = now
            start_date = now - timedelta(days=months * 30)
            return {
                "start": start_date.strftime("%Y-%m-%d"),
                "end": end_date.strftime("%Y-%m-%d")
            }
        
        # "최근 N일", "최근 N일간"
        days_match = re.search(r'최근\s*(\d+)\s*일', question_lower)
        if days_match:
            days = int(days_match.group(1))
            end_date = now
            start_date = now - timedelta(days=days)
            return {
                "start": start_date.strftime("%Y-%m-%d"),
                "end": end_date.strftime("%Y-%m-%d")
            }
        
        # "월별", "월별 통계", "월별 현황" -> 최근 12개월
        if any(keyword in question_lower for keyword in ["월별", "월별 통계", "월별 현황", "월별 추이"]):
            end_date = now
            start_date = now - timedelta(days=365)  # 최근 12개월
            return {
                "start": start_date.strftime("%Y-%m-%d"),
                "end": end_date.strftime("%Y-%m-%d")
            }
        
        
        
        # 기본값: 날짜 범위 없음
        return None
    
    async def _call_ollama(self, prompt: str) -> str:
        """Ollama API 호출 헬퍼 메서드"""
        async with httpx.AsyncClient() as client:
            try:
                response = await client.post(
                    self.url,
                    json={
                        "model": self.model,
                        "prompt": prompt,
                        "stream": False
                    },
                    timeout=30.0
                )
                result = response.json()
                return result.get("response", "").strip()
            except Exception as e:
                logger.error(f"❌ Ollama API 호출 실패: {str(e)}")
                raise
    
    def _parse_json_response(self, generated_text: str) -> Optional[Union[dict, list, str]]:
        """AI 응답에서 JSON 추출"""
        generated_text = generated_text.strip()
        
        # Try to decode the first JSON object in the model output robustly
        decoder = json.JSONDecoder()
        query_data = None
        
        try:
            obj, idx = decoder.raw_decode(generated_text)
            query_data = obj
        except ValueError:
            # fallback: try from first '{' or '['
            start = generated_text.find("{")
            if start == -1:
                start = generated_text.find("[")
            if start != -1:
                try:
                    obj, idx = decoder.raw_decode(generated_text[start:])
                    query_data = obj
                except Exception:
                    query_data = None

        # Additional heuristics: regex extract first {...} or [...] (DOTALL)
        if not query_data:
            try:
                m = re.search(r"(\{.*\}|\[.*\])", generated_text, re.DOTALL)
                if m:
                    candidate = m.group(1)
                    query_data = json.loads(candidate)
            except Exception:
                query_data = None

        # Try to repair common issues (single quotes, trailing commas)
        if not query_data:
            try:
                repaired = generated_text.replace("'", '"')
                repaired = re.sub(r",(\s*[}\]])", r"\1", repaired)
                m = re.search(r"(\{.*\}|\[.*\])", repaired, re.DOTALL)
                if m:
                    query_data = json.loads(m.group(1))
            except Exception:
                query_data = None
        
        return query_data
    
    async def _extract_filters(self, question: str) -> Dict[str, Any]:
        """1단계: 필터 값 추출"""
        logger.info("🔍 1단계: 필터 추출 중...")
        
        prompt = f"""{FILTER_SELECT_PROMPT}
질문: {question}
"""
        
        try:
            response = await self._call_ollama(prompt)
            filters = self._parse_json_response(response)
            
            if isinstance(filters, dict):
                logger.info(f"✅ 필터 추출 완료: {list(filters.keys())}")
                return filters
            else:
                logger.warning("⚠️ 필터 추출 실패, 빈 필터 반환")
                return {}
        except Exception as e:
            logger.error(f"❌ 필터 추출 실패: {str(e)}")
            return {}
    
    async def _select_tables(self, question: str) -> Union[str, List[str]]:
        """2단계: 테이블 유형 선택"""
        logger.info("📋 2단계: 테이블 선택 중...")
        
        prompt = TABLE_SELECT_PROMPT.replace("{question}", question)
        
        try:
            response = await self._call_ollama(prompt)
            response = response.strip()
            logger.debug(f"테이블 선택 응답: {response}")
            
            # 응답에서 JSON 추출 시도
            table_data = self._parse_json_response(response)
            if isinstance(table_data, list):
                # 리스트 형식
                logger.info(f"✅ 테이블 선택 완료: {table_data}")
                return table_data
            elif isinstance(table_data, str):
                # 문자열 형식
                if table_data in ["class", "discharge"]:
                    logger.info(f"✅ 테이블 선택 완료: {table_data}")
                    return table_data
            
            # 직접 문자열 매칭
            response_lower = response.lower()
            
            # "class", "discharge" 단일 테이블
            if response_lower == "class" or (response_lower.startswith("class") and "discharge" not in response_lower):
                logger.info("✅ 테이블 선택 완료: class")
                return "class"
            elif response_lower == "discharge" or (response_lower.startswith("discharge") and "class" not in response_lower):
                logger.info("✅ 테이블 선택 완료: discharge")
                return "discharge"
            
            # "class"와 "discharge" 둘 다 포함된 경우
            if "class" in response_lower and "discharge" in response_lower:
                logger.info("✅ 테이블 선택 완료: [class, discharge]")
                return ["class", "discharge"]
            
            # 리스트 형식 문자열 파싱 시도
            if "[" in response and "]" in response:
                # "[class, discharge]" 형식 추출
                list_match = re.search(r'\[([^\]]+)\]', response)
                if list_match:
                    items = [item.strip().strip('"\'') for item in list_match.group(1).split(",")]
                    valid_items = [item for item in items if item in ["class", "discharge"]]
                    if len(valid_items) == 2:
                        logger.info("✅ 테이블 선택 완료: [class, discharge]")
                        return ["class", "discharge"]
                    elif len(valid_items) == 1:
                        logger.info(f"✅ 테이블 선택 완료: {valid_items[0]}")
                        return valid_items[0]
            
            # 질문 내용 기반 추론
            question_lower = question.lower()
            if any(keyword in question_lower for keyword in ["입퇴소", "입소.*퇴소", "퇴소.*입소"]):
                logger.info("✅ 질문 기반 추론: [class, discharge]")
                return ["class", "discharge"]
            elif any(keyword in question_lower for keyword in ["퇴소", "퇴원", "사유", "증감"]):
                logger.info("✅ 질문 기반 추론: discharge")
                return "discharge"
            elif any(keyword in question_lower for keyword in ["입소", "재원", "현재 학생", "담당 학생"]):
                logger.info("✅ 질문 기반 추론: class")
                return "class"
            
            # 기본값: class
            logger.warning(f"⚠️ 테이블 선택 실패 (응답: {response[:100]}), 기본값 'class' 반환")
            return "class"
        except Exception as e:
            logger.error(f"❌ 테이블 선택 실패: {str(e)}")
            # 예외 발생 시 질문 기반 추론
            question_lower = question.lower()
            if any(keyword in question_lower for keyword in ["입퇴소"]):
                return ["class", "discharge"]
            elif any(keyword in question_lower for keyword in ["퇴소", "퇴원"]):
                return "discharge"
            return "class"
    
    def _extract_columns_from_question(self, question: str, table_type: str) -> List[str]:
        """질문에서 필요한 컬럼 추출"""
        question_lower = question.lower()
        columns = []
        
        # 기본 컬럼 (항상 포함)
        base_columns = ["student_name", "start_date", "grade", "class_name"]
        
        # 테이블별 기본 컬럼
        if table_type == "class":
            base_columns.extend(["student_name", "start_date", "grade", "class_name", "parent_phone_number"])
        elif table_type == "discharge":
            base_columns.extend(["student_name", "grade", "class_name", "discharge_date", "start_date", "discharging_reason", "parent_phone_number"])
        
        # 질문에서 명시적으로 언급된 컬럼 확인
        column_keywords = {
            "student_name": ["학생명", "학생", "이름"],
            "teacher_name": ["선생님", "담당", "원장"],
            "class_name": ["반", "수업", "과목"],
            "grade": ["학년"],
            "school_name": ["학교"],
            "start_date": ["입소일", "입소일자", "시작일"],
            "discharge_date": ["퇴소일", "퇴소일자", "퇴원일"],
            "discharging_reason": ["사유", "퇴원사유", "퇴소사유"],
            "parent_phone_number": ["전화", "연락처", "학부모"]
        }
        
        # 질문에서 언급된 컬럼 추가
        for col, keywords in column_keywords.items():
            if any(keyword in question_lower for keyword in keywords):
                if col not in columns:
                    columns.append(col)
        
        # 기본 컬럼 추가 (중복 제거)
        for col in base_columns:
            if col not in columns:
                columns.append(col)
        
        return columns
    
    def _determine_sort_by(self, question: str, table_type: str) -> Optional[str]:
        """정렬 기준 결정"""
        question_lower = question.lower()
        
        if table_type == "class":
            if any(keyword in question_lower for keyword in ["월별", "통계", "현황", "추이"]):
                return "start_date"
            return "start_date"
        elif table_type == "discharge":
            if any(keyword in question_lower for keyword in ["월별", "통계", "현황", "추이"]):
                return "discharge_date"
            return "discharge_date"
        
        return None
    
    def _determine_aggregations(self, question: str) -> Optional[List[str]]:
        """집계 함수 결정"""
        question_lower = question.lower()
        
        if any(keyword in question_lower for keyword in ["월별", "통계", "현황", "추이", "요약"]):
            return ["count_by_month"]
        
        return None
    
    def _generate_json_query(self, question: str, filters: Dict[str, Any], table_type: Union[str, List[str]]) -> Union[dict, list]:
        """3단계: 최종 JSON 쿼리 생성 (로직 처리)"""
        logger.info("📝 3단계: JSON 쿼리 생성 중...")
        
        # table_type이 리스트인 경우 각 테이블에 대해 쿼리 생성
        if isinstance(table_type, list):
            queries = []
            for table in table_type:
                columns = self._extract_columns_from_question(question, table)
                sort_by = self._determine_sort_by(question, table)
                aggregations = self._determine_aggregations(question)
                
                query_obj = {
                    "target_table": table,
                    "filters": filters,
                    "columns": columns,
                    "aggregations": aggregations,
                    "sort_by": sort_by,
                    "date_range": None  # 로직에서 나중에 설정
                }
                queries.append(query_obj)
            
            logger.info(f"✅ JSON 쿼리 생성 완료: {len(queries)}개 쿼리")
            return queries
        
        # table_type이 문자열인 경우 단일 쿼리 생성
        else:
            columns = self._extract_columns_from_question(question, table_type)
            sort_by = self._determine_sort_by(question, table_type)
            aggregations = self._determine_aggregations(question)
            
            query_obj = {
                "target_table": table_type,
                "filters": filters,
                "columns": columns,
                "aggregations": aggregations,
                "sort_by": sort_by,
                "date_range": None  # 로직에서 나중에 설정
            }
            
            logger.info("✅ JSON 쿼리 생성 완료")
            return query_obj
    
    async def analyze_question(self, question: str) -> Union[ReportQuery, List[ReportQuery]]:
        """자연어 질문을 구조화된 쿼리로 변환 (3단계 프로세스)"""
        logger.info(f"🤖 AI 분석 시작: {question[:50]}...")
        
        try:
            # 1단계: 필터 추출
            filters = await self._extract_filters(question)
            
            # 2단계: 테이블 선택
            table_type = await self._select_tables(question)
            
            # 3단계: JSON 쿼리 생성 (로직 처리, await 불필요)
            query_data = self._generate_json_query(question, filters, table_type)
            
            if not query_data or not isinstance(query_data, (dict, list)):
                logger.error("AI output (for debugging): %s", query_data)
                raise ValueError("JSON not found or invalid in model output")

            # 날짜 범위를 로직으로 계산
            parsed_date_range = self._parse_date_range(question)
            
            # If model returned a list of query objects, convert to list of ReportQuery
            if isinstance(query_data, list):
                queries = []
                for item in query_data:
                    if not isinstance(item, dict):
                        continue
                    target_table = item.get("target_table") or (item.get("target_tables", [None])[0] if isinstance(item.get("target_tables"), list) else item.get("target_tables"))
                    
                    # 테이블 타입에 따라 날짜 속성 설정
                    date_range_with_property = None
                    if parsed_date_range:
                        date_range_with_property = parsed_date_range.copy()
                        if target_table == "class":
                            date_range_with_property["property"] = "start_date"
                        elif target_table == "discharge":
                            date_range_with_property["property"] = "discharge_date"
                    
                    q = ReportQuery(
                        target_table=target_table,
                        filters=item.get("filters", {}),
                        columns=item.get("columns", []),
                        aggregations=item.get("aggregations"),
                        date_range=date_range_with_property
                    )
                    queries.append(q)
                logger.info(f"✅ 쿼리 생성 완료: {', '.join([str(q.target_table) for q in queries])} 테이블")
                if parsed_date_range:
                    for q in queries:
                        if q.date_range:
                            logger.info(f"📅 {q.target_table} 테이블 날짜 범위 ({q.date_range.get('property')}): {q.date_range['start']} ~ {q.date_range['end']}")
                return queries

            # Normalize target table(s) for single-object response
            target_table = query_data.get("target_table")
            if not target_table:
                tts = query_data.get("target_tables")
                if isinstance(tts, list) and tts:
                    target_table = tts[0]
                elif isinstance(tts, str):
                    target_table = tts

            # 테이블 타입에 따라 날짜 속성 설정
            date_range_with_property = None
            if parsed_date_range:
                date_range_with_property = parsed_date_range.copy()
                if target_table == "class":
                    date_range_with_property["property"] = "start_date"
                elif target_table == "discharge":
                    date_range_with_property["property"] = "discharge_date"

            query = ReportQuery(
                target_table=target_table,
                filters=query_data.get("filters", {}),
                columns=query_data.get("columns", []),
                aggregations=query_data.get("aggregations"),
                date_range=date_range_with_property
            )

            logger.info(f"✅ 쿼리 생성 완료: {query.target_table} 테이블")
            if parsed_date_range:
                if query.date_range:
                    logger.info(f"📅 날짜 범위 ({query.date_range.get('property')}): {query.date_range['start']} ~ {query.date_range['end']}")
            print(query)
            return query
                
        except Exception as e:
            logger.error(f"❌ AI 분석 실패: {str(e)}")
            # 기본 쿼리 반환
            return ReportQuery(
                target_table="class",
                columns=["학생명", "담당", "반명"]
            )


####
class ExcelFileHandler:
    """input 폴더의 엑셀 파일을 감지하고 폴더별로 구별해서 저장하는 클래스"""
    
    def __init__(self, notion_manager: Optional[NotionManager] = None):
        self.input_dir = Path("input")
        self.processed_files = set()  # 처리된 파일 추적 (중복 방지)
        self.queued_files = set()  # 큐에 추가된 파일 추적 (중복 큐 추가 방지)
        self.table_folders = {
            "class": self.input_dir / "class",
            "discharge": self.input_dir / "discharge",
            "student": self.input_dir / "student"
        }
        # 각 폴더별로 읽은 파일들을 저장
        self.stored_files = {
            "class": [],
            "discharge": [],
            "student": []
        }
        # 전처리 필터 키워드 (반명에 포함되면 제거)
        self.filter_keywords = ["TEST", "면접", "자소서", "상담", "대입"]
        # NotionManager (날짜 범위 조회용)
        self.notion = notion_manager
    
    def _read_excel_file(self, file_path: Path) -> Optional[pd.DataFrame]:
        """엑셀 파일을 읽어서 DataFrame으로 반환 (빈 행 제외)"""
        try:
            logger.info(f"📖 엑셀 파일 읽기 시작: {file_path.name}")
            
            # 엑셀 파일 읽기
            df = pd.read_excel(file_path)
            
            if df.empty:
                logger.warning(f"⚠️ 빈 엑셀 파일: {file_path.name}")
                return None
            
            # 빈 행 제거 (모든 컬럼이 NaN인 행)
            before_count = len(df)
            df = df.dropna(how='all')  # 모든 값이 NaN인 행 제거
            
            if len(df) < before_count:
                logger.info(f"🗑️ 빈 행 {before_count - len(df)}개 제거됨")
            
            # 빈 열 제거 (모든 값이 NaN인 열)
            df = df.dropna(axis=1, how='all')
            
            if df.empty:
                logger.warning(f"⚠️ 빈 행 제거 후 데이터가 없음: {file_path.name}")
                return None
            
            logger.info(f"✅ 엑셀 파일 읽기 완료: {file_path.name} ({len(df)}개 행)")
            return df
            
        except Exception as e:
            logger.error(f"❌ 엑셀 파일 읽기 실패: {file_path.name}, 오류: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return None
    
    def watch_and_store(self) -> Dict[str, List[Dict[str, Any]]]:
        """input 폴더를 스캔하여 새 엑셀 파일을 감지하고 폴더별로 구별해서 저장
        
        Returns:
            Dict[str, List[Dict]]: 폴더별로 저장된 파일 정보
                예: {
                    "class": [
                        {"file_name": "class1.xlsx", "file_path": "input/class/class1.xlsx", "dataframe": df1},
                        {"file_name": "class2.xlsx", "file_path": "input/class/class2.xlsx", "dataframe": df2}
                    ],
                    "discharge": [
                        {"file_name": "discharge1.xlsx", "file_path": "input/discharge/discharge1.xlsx", "dataframe": df3}
                    ],
                    "student": []
                }
        """
        # input 폴더가 없으면 생성
        if not self.input_dir.exists():
            self.input_dir.mkdir(parents=True, exist_ok=True)
            logger.info(f"📁 input 폴더 생성: {self.input_dir}")
        
        new_files_count = 0
        
        # 각 테이블별 폴더 확인 및 파일 읽기
        for table_type, folder_path in self.table_folders.items():
            logger.info(f"📂 [{table_type}] 폴더 스캔 중: {folder_path}")
            
            if not folder_path.exists():
                folder_path.mkdir(parents=True, exist_ok=True)
                logger.info(f"📁 {table_type} 폴더 생성: {folder_path}")
                continue
            
            # 엑셀 파일 찾기
            excel_files = list(folder_path.glob("*.xlsx")) + list(folder_path.glob("*.xls"))
            logger.info(f"📋 [{table_type}] 폴더에서 {len(excel_files)}개 엑셀 파일 발견")
            
            for excel_file in excel_files:
                # 파일 경로를 키로 사용하여 처리 여부 확인
                file_key = str(excel_file.resolve())
                
                # 이미 처리되었거나 큐에 추가된 파일은 건너뜀
                if file_key in self.processed_files:
                    logger.debug(f"⏭️ [{table_type}] 이미 처리된 파일 건너뜀: {excel_file.name}")
                    continue
                
                if file_key in self.queued_files:
                    logger.debug(f"⏭️ [{table_type}] 이미 큐에 추가된 파일 건너뜀: {excel_file.name}")
                    continue
                
                # 이미 stored_files에 있는 파일인지 확인 (같은 파일이 여러 번 스캔되는 것 방지)
                already_stored = False
                for stored_file in self.stored_files[table_type]:
                    if stored_file.get("file_key") == file_key:
                        already_stored = True
                        break
                
                if already_stored:
                    logger.debug(f"⏭️ [{table_type}] 이미 저장된 파일 건너뜀: {excel_file.name}")
                    continue
                
                # 엑셀 파일 읽기
                df = self._read_excel_file(excel_file)
                
                if df is not None:
                    # 상대 경로 생성 (안전하게)
                    try:
                        file_path_str = str(excel_file.resolve().relative_to(Path.cwd().resolve()))
                    except ValueError:
                        # 상대 경로 변환 실패 시 절대 경로 사용
                        file_path_str = str(excel_file.resolve())
                    
                    file_info = {
                        "file_name": excel_file.name,
                        "file_path": file_path_str,
                        "file_key": file_key,  # 파일 키 추가
                        "folder": table_type,
                        "dataframe": df,
                        "rows": len(df),
                        "columns": list(df.columns),
                        "read_time": datetime.now().isoformat()
                    }
                    # 폴더별로 구별해서 저장
                    self.stored_files[table_type].append(file_info)
                    # 파일을 읽은 즉시 processed_files에 추가하여 다음 스캔에서 건너뛰도록 함
                    # (파일 이동 후에도 다시 추가되지만, 이미 processed_files에 있으면 건너뜀)
                    self.processed_files.add(file_key)
                    new_files_count += 1
                    logger.info(f"✅ [{table_type}] 파일 저장 완료: {excel_file.name} ({len(df)}개 행)")
                else:
                    logger.warning(f"⚠️ [{table_type}] 파일 읽기 실패: {excel_file.name}")
        
        # 폴더별 요약 로그
        for table_type, files in self.stored_files.items():
            if files:
                logger.info(f"📊 [{table_type}] 폴더: 총 {len(files)}개 파일 저장됨")
        
        if new_files_count > 0:
            logger.info(f"🎉 새로 감지된 파일: {new_files_count}개")
        else:
            logger.info("💤 새로운 파일이 없습니다.")
        
        return self.stored_files
    
    async def preprocess_and_merge(self, table_type: str) -> Optional[pd.DataFrame]:
        """저장된 파일들을 합치고 중복 제거 및 날짜 필터링
        
        Args:
            table_type: 테이블 타입 (class, discharge, student)
        
        Returns:
            전처리된 DataFrame (합쳐지고 중복 제거, 날짜 필터링됨) 또는 None
        """
        if table_type not in self.stored_files:
            logger.error(f"❌ 알 수 없는 테이블 타입: {table_type}")
            return None
        
        file_list = self.stored_files[table_type]
        
        if not file_list:
            logger.warning(f"⚠️ [{table_type}] 폴더에 저장된 파일이 없습니다.")
            return None
        
        logger.info(f"🔄 [{table_type}] 전처리 시작: {len(file_list)}개 파일 합치기")
        
        # 모든 DataFrame 합치기
        dataframes = []
        for file_info in file_list:
            df = file_info.get("dataframe")
            if df is not None and not df.empty:
                # 원본 파일 정보를 컬럼으로 추가 (선택적)
                df_copy = df.copy()
                dataframes.append(df_copy)
                logger.debug(f"  - {file_info['file_name']}: {len(df)}개 행 추가")
        
        if not dataframes:
            logger.warning(f"⚠️ [{table_type}] 합칠 수 있는 데이터가 없습니다.")
            return None
        
        # 모든 DataFrame 합치기
        merged_df = pd.concat(dataframes, ignore_index=True)
        original_count = len(merged_df)
        logger.info(f"📊 [{table_type}] 합친 데이터: {original_count}개 행")
        
        # 중복 제거
        # 모든 컬럼이 동일한 행을 중복으로 간주
        deduplicated_df = merged_df.drop_duplicates(keep='first')
        removed_count = original_count - len(deduplicated_df)
        
        if removed_count > 0:
            logger.info(f"🗑️ [{table_type}] 중복 제거: {removed_count}개 행 제거됨 ({original_count} → {len(deduplicated_df)})")
        else:
            logger.info(f"✅ [{table_type}] 중복 데이터 없음")
        
        # 반명 필터링 (TEST, 면접, 자소서, 상담, 대입 포함된 데이터 제거)
        before_filter_count = len(deduplicated_df)
        filtered_df = self._filter_by_class_name(deduplicated_df)
        filter_removed_count = before_filter_count - len(filtered_df)
        
        if filter_removed_count > 0:
            logger.info(f"🔍 [{table_type}] 반명 필터링: {filter_removed_count}개 행 제거됨 (필터 키워드: {self.filter_keywords})")
        else:
            logger.info(f"✅ [{table_type}] 반명 필터링: 제거된 데이터 없음")
        
        # 날짜 필터링 (Notion에 있는 날짜 범위 제외)
        before_date_filter_count = len(filtered_df)
        date_filtered_df = await self._filter_by_notion_date_range(filtered_df, table_type)
        #date_filtered_df = filtered_df
        date_filter_removed_count = before_date_filter_count - len(date_filtered_df)
        
        if date_filter_removed_count > 0:
            logger.info(f"📅 [{table_type}] 날짜 필터링: {date_filter_removed_count}개 행 제거됨 (Notion 날짜 범위 제외)")
        else:
            logger.info(f"✅ [{table_type}] 날짜 필터링: 제거된 데이터 없음")
        
        logger.info(f"✅ [{table_type}] 전처리 완료: 최종 {len(date_filtered_df)}개 행 (원본: {original_count} → 중복제거: {before_filter_count} → 반명필터: {before_date_filter_count} → 날짜필터: {len(date_filtered_df)})")
        
        return date_filtered_df
    
    async def _filter_by_notion_date_range(self, df: pd.DataFrame, table_type: str) -> pd.DataFrame:
        """Notion의 날짜 범위에 포함된 데이터 제거
        
        Args:
            df: 필터링할 DataFrame
            table_type: 테이블 타입 (class, discharge, student)
        
        Returns:
            날짜 범위 밖의 데이터만 남은 DataFrame
        """
        if df.empty:
            return df
        
        if not self.notion:
            logger.warning("⚠️ NotionManager가 설정되지 않아 날짜 필터링을 건너뜁니다.")
            return df
        
        # 날짜 속성명 결정
        date_property = None
        if table_type in ["class", "student"]:
            date_property = "start_date"
        elif table_type == "discharge":
            date_property = "discharge_date"
        else:
            logger.warning(f"⚠️ 알 수 없는 테이블 타입: {table_type}, 날짜 필터링 건너뜀")
            return df
        
        # Notion에서 날짜 범위 조회
        date_range = await self.notion.get_date_range_from_table(table_type, date_property)
        
        if not date_range:
            logger.info(f"📭 [{table_type}] Notion에 데이터가 없어 전체 데이터를 유지합니다.")
            return df
        
        first_date, last_date = date_range
        logger.info(f"📅 [{table_type}] Notion 날짜 범위: {first_date.date()} ~ {last_date.date()}")
        
        # 날짜 컬럼 찾기 (더 유연하게)
        date_col = None
        for col in df.columns:
            col_str = str(col).strip()
            col_lower = col_str.lower()
            
            if date_property == "start_date":
                # 입소일자 관련 키워드
                if any(keyword in col_str for keyword in ['입소일자', '입소일', '입소 날짜', '시작일', '시작 날짜']):
                    date_col = col
                    break
                elif 'start_date' in col_lower or 'startdate' in col_lower:
                    date_col = col
                    break
                elif '날짜' in col_str and any(keyword in col_str for keyword in ['입소', '시작']):
                    date_col = col
                    break
            elif date_property == "discharge_date":
                # 퇴소일자 관련 키워드
                if any(keyword in col_str for keyword in ['퇴소일자', '퇴소일', '퇴원일자', '퇴원일', '퇴소 날짜', '퇴원 날짜']):
                    date_col = col
                    break
                elif 'discharge_date' in col_lower or 'dischargedate' in col_lower:
                    date_col = col
                    break
                elif '날짜' in col_str and any(keyword in col_str for keyword in ['퇴소', '퇴원']):
                    date_col = col
                    break
        
        if not date_col:
            # 디버깅: 사용 가능한 컬럼명 출력
            available_cols = [str(col) for col in df.columns]
            logger.warning(f"⚠️ [{table_type}] 날짜 컬럼을 찾을 수 없어 날짜 필터링을 건너뜁니다.")
            logger.debug(f"   사용 가능한 컬럼: {available_cols}")
            logger.debug(f"   찾는 날짜 속성: {date_property}")
            return df
        
        # 날짜 컬럼을 datetime으로 변환
        try:
            df[date_col] = pd.to_datetime(df[date_col])
        except Exception as e:
            logger.error(f"❌ 날짜 컬럼 변환 실패: {e}")
            return df
        
        # 날짜 범위 밖의 데이터만 남기기 (범위 내 데이터 제거)
        # first_date <= 날짜 <= last_date 범위의 데이터 제거
        before_count = len(df)
        filtered_df = df[(df[date_col] < first_date) | (df[date_col] > last_date)]
        removed_count = before_count - len(filtered_df)
        
        if removed_count > 0:
            logger.info(f"🗑️ [{table_type}] 날짜 범위 내 데이터 {removed_count}개 제거됨 ({first_date.date()} ~ {last_date.date()})")
        
        return filtered_df.reset_index(drop=True)
    
    def _filter_by_class_name(self, df: pd.DataFrame) -> pd.DataFrame:
        """반명에 필터 키워드가 포함된 행 제거
        
        Args:
            df: 전처리할 DataFrame
        
        Returns:
            필터링된 DataFrame
        """
        if df.empty:
            return df
        
        # 반명 컬럼 찾기 (유연하게)
        class_name_col = None
        for col in df.columns:
            col_lower = str(col).lower()
            if '반명' in col_lower or 'class_name' in col_lower or '반' in col_lower:
                class_name_col = col
                break
        
        if not class_name_col:
            logger.warning("⚠️ 반명 컬럼을 찾을 수 없어 필터링을 건너뜁니다.")
            return df
        
        # 필터링: 반명에 키워드가 포함된 행 제거
        def should_filter_row(class_name_value):
            if pd.isna(class_name_value):
                return False
            class_name_upper = str(class_name_value).upper()
            for keyword in self.filter_keywords:
                if keyword in class_name_upper:
                    return True
            return False
        
        filtered_df = df[~df[class_name_col].apply(should_filter_row)]
        
        return filtered_df.reset_index(drop=True)
    
    async def preprocess_all_folders(self) -> Dict[str, Optional[pd.DataFrame]]:
        """모든 폴더의 파일들을 전처리 (합치기 + 중복 제거 + 날짜 필터링)
        
        Returns:
            Dict[str, Optional[pd.DataFrame]]: 폴더별 전처리된 DataFrame
        """
        result = {}
        
        for table_type in ["class", "discharge", "student"]:
            result[table_type] = await self.preprocess_and_merge(table_type)
        
        return result
    
    def get_stored_files(self, table_type: Optional[str] = None) -> Dict[str, List[Dict[str, Any]]]:
        """저장된 파일 정보 조회
        
        Args:
            table_type: 특정 테이블 타입만 조회 (None이면 전체)
        
        Returns:
            폴더별로 저장된 파일 정보
        """
        if table_type:
            return {table_type: self.stored_files.get(table_type, [])}
        return self.stored_files.copy()
    
    def clear_stored_files(self, table_type: Optional[str] = None):
        """저장된 파일 정보 초기화
        
        Args:
            table_type: 특정 테이블 타입만 초기화 (None이면 전체)
        """
        if table_type:
            if table_type in self.stored_files:
                self.stored_files[table_type] = []
                logger.info(f"🔄 [{table_type}] 폴더의 저장된 파일 정보 초기화 완료")
        else:
            for table_type in self.stored_files:
                self.stored_files[table_type] = []
            logger.info("🔄 모든 폴더의 저장된 파일 정보 초기화 완료")
    
    def reset_processed_files(self):
        """처리된 파일 목록 초기화 (모든 파일을 다시 읽을 수 있도록)"""
        self.processed_files.clear()
        self.queued_files.clear()  # 큐 목록도 함께 초기화
        logger.info("🔄 처리된 파일 목록 및 큐 목록 초기화 완료")
    
    def move_processed_files_to_imported(self, table_type: str) -> int:
        """처리된 파일들을 imported 폴더로 이동
        
        Args:
            table_type: 테이블 타입 (class, discharge, student)
        
        Returns:
            이동된 파일 수
        """
        if table_type not in self.stored_files:
            logger.error(f"❌ 알 수 없는 테이블 타입: {table_type}")
            return 0
        
        file_list = self.stored_files[table_type]
        if not file_list:
            logger.warning(f"⚠️ [{table_type}] 이동할 파일이 없습니다.")
            return 0
        
        # imported 폴더 생성
        imported_dir = self.input_dir / "imported" / table_type
        imported_dir.mkdir(parents=True, exist_ok=True)
        
        moved_count = 0
        
        for file_info in file_list:
            try:
                file_path = Path(file_info["file_path"])
                file_key = file_info.get("file_key")
                
                # 절대 경로로 변환
                if not file_path.is_absolute():
                    file_path = Path.cwd() / file_path
                
                if not file_path.exists():
                    logger.warning(f"⚠️ 파일이 존재하지 않음: {file_path}")
                    # 파일이 없어도 추적 목록에서 제거
                    if file_key:
                        self.queued_files.discard(file_key)
                        self.processed_files.add(file_key)
                    continue
                
                # imported 폴더로 이동할 파일명 생성 (타임스탬프 추가로 중복 방지)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name = file_path.name
                name_parts = file_name.rsplit('.', 1)
                if len(name_parts) == 2:
                    new_file_name = f"{name_parts[0]}_{timestamp}.{name_parts[1]}"
                else:
                    new_file_name = f"{file_name}_{timestamp}"
                
                dest_path = imported_dir / new_file_name
                
                # 파일 이동
                file_path.rename(dest_path)
                moved_count += 1
                logger.info(f"📦 [{table_type}] 파일 이동: {file_path.name} → {dest_path}")
                
                # 파일 이동 성공 시 추적 목록 업데이트
                if file_key:
                    self.queued_files.discard(file_key)  # 큐 목록에서 제거
                    self.processed_files.add(file_key)  # 처리 완료 목록에 추가
                
            except Exception as e:
                logger.error(f"❌ [{table_type}] 파일 이동 실패: {file_info['file_name']}, 오류: {e}")
        
        if moved_count > 0:
            logger.info(f"✅ [{table_type}] {moved_count}개 파일을 imported 폴더로 이동 완료")
            # 이동된 파일은 stored_files에서 제거
            self.stored_files[table_type] = []
        
        return moved_count
    
    def move_specific_files_to_imported(self, table_type: str, file_infos: List[Dict[str, Any]]) -> int:
        """특정 파일들만 imported 폴더로 이동
        
        Args:
            table_type: 테이블 타입 (class, discharge, student)
            file_infos: 이동할 파일 정보 목록
        
        Returns:
            이동된 파일 수
        """
        if table_type not in self.stored_files:
            logger.error(f"❌ 알 수 없는 테이블 타입: {table_type}")
            return 0
        
        if not file_infos:
            logger.warning(f"⚠️ [{table_type}] 이동할 파일 정보가 없습니다.")
            return 0
        
        # imported 폴더 생성
        imported_dir = self.input_dir / "imported" / table_type
        imported_dir.mkdir(parents=True, exist_ok=True)
        
        moved_count = 0
        moved_file_keys = set()
        
        for file_info in file_infos:
            try:
                file_path = Path(file_info["file_path"])
                file_key = file_info.get("file_key")
                
                # 절대 경로로 변환
                if not file_path.is_absolute():
                    file_path = Path.cwd() / file_path
                
                if not file_path.exists():
                    logger.warning(f"⚠️ 파일이 존재하지 않음: {file_path}")
                    # 파일이 없어도 추적 목록에서 제거
                    if file_key:
                        self.queued_files.discard(file_key)
                        self.processed_files.add(file_key)
                        moved_file_keys.add(file_key)
                    continue
                
                # imported 폴더로 이동할 파일명 생성 (타임스탬프 추가로 중복 방지)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name = file_path.name
                name_parts = file_name.rsplit('.', 1)
                if len(name_parts) == 2:
                    new_file_name = f"{name_parts[0]}_{timestamp}.{name_parts[1]}"
                else:
                    new_file_name = f"{file_name}_{timestamp}"
                
                dest_path = imported_dir / new_file_name
                
                # 파일 이동
                file_path.rename(dest_path)
                moved_count += 1
                logger.info(f"📦 [{table_type}] 파일 이동: {file_path.name} → {dest_path}")
                
                # 파일 이동 성공 시 추적 목록 업데이트
                if file_key:
                    self.queued_files.discard(file_key)  # 큐 목록에서 제거
                    self.processed_files.add(file_key)  # 처리 완료 목록에 추가
                    moved_file_keys.add(file_key)
                
            except Exception as e:
                logger.error(f"❌ [{table_type}] 파일 이동 실패: {file_info.get('file_name', 'unknown')}, 오류: {e}")
        
        # stored_files에서 이동된 파일들만 제거
        if moved_file_keys:
            remaining_files = [
                f for f in self.stored_files[table_type]
                if f.get("file_key") not in moved_file_keys
            ]
            self.stored_files[table_type] = remaining_files
            logger.debug(f"🔄 [{table_type}] stored_files에서 {len(moved_file_keys)}개 파일 제거, {len(remaining_files)}개 파일 남음")
        
        if moved_count > 0:
            logger.info(f"✅ [{table_type}] {moved_count}개 파일을 imported 폴더로 이동 완료")
        
        return moved_count

####

class ExcelImporter:
    def __init__(self, notion_manager):
        self.notion = notion_manager

    async def get_date_range_from_notion(self, table_type: str, date_property: str) -> Optional[datetime]:
        """노션 DB에서 마지막 날짜 조회
        
        Args:
            table_type: 테이블 타입 (class, discharge, student)
            date_property: 날짜 속성명 (start_date, discharge_date 등)
        
        Returns:
            마지막 날짜 (datetime) 또는 None
        """
        try:
            db_id = self.notion.db_map.get(table_type.lower())
            if not db_id:
                logger.error(f"❌ 테이블을 찾을 수 없음: {table_type}")
                return None
            
            # 가장 최신 데이터만 조회
            newest = await self.notion.client.databases.query(
                database_id=db_id,
                sorts=[{"property": date_property, "direction": "descending"}],
                page_size=1
            )
            
            if newest.get('results'):
                date_value = self.notion._get_date(newest['results'][0], date_property)
                if date_value:
                    try:
                        if "T" in date_value:
                            return datetime.fromisoformat(date_value.split("T")[0])
                        else:
                            return datetime.fromisoformat(date_value)
                    except Exception as e:
                        logger.error(f"❌ 날짜 파싱 실패: {e}")
                        return None
            
            return None
        except Exception as e:
            logger.error(f"❌ 노션 조회 실패: {e}")
            return None

    
    def _convert_dataframe_row_to_notion_properties(self, row: pd.Series, table_type: str, df: pd.DataFrame) -> Dict[str, Any]:
        """DataFrame 행을 Notion 속성 형식으로 변환
        
        Args:
            row: DataFrame의 한 행
            table_type: 테이블 타입 (class, discharge, student)
            df: 전체 DataFrame (컬럼 정보 확인용)
        
        Returns:
            Notion 속성 딕셔너리
        """
        properties = {}
        
        # 테이블 타입별 매핑
        if table_type == "class":
            # class 테이블 속성 매핑
            if "학생명" in df.columns or "student_name" in df.columns:
                col = "학생명" if "학생명" in df.columns else "student_name"
                student_name = str(row[col]) if pd.notna(row[col]) else ""
                properties["student_name"] = {"title": [{"text": {"content": student_name}}]}
            
            if "담당" in df.columns or "teacher_name" in df.columns:
                col = "담당" if "담당" in df.columns else "teacher_name"
                teacher_name = row[col]
                if pd.notna(teacher_name):
                    if isinstance(teacher_name, str):
                        # 쉼표로 구분된 경우 리스트로 변환
                        if "," in teacher_name:
                            teacher_list = [t.strip() for t in teacher_name.split(",")]
                            properties["teacher_name"] = {"multi_select": [{"name": str(t)} for t in teacher_list]}
                        else:
                            properties["teacher_name"] = {"rich_text": [{"text": {"content": str(teacher_name)}}]}
                    elif isinstance(teacher_name, list):
                        properties["teacher_name"] = {"multi_select": [{"name": str(t)} for t in teacher_name]}
            
            if "반명" in df.columns or "class_name" in df.columns:
                col = "반명" if "반명" in df.columns else "class_name"
                class_name = str(row[col]) if pd.notna(row[col]) else ""
                properties["class_name"] = {"rich_text": [{"text": {"content": class_name}}]}
            
            if "부모HP" in df.columns or "parent_phone_number" in df.columns:
                col = "부모HP" if "부모HP" in df.columns else "parent_phone_number"
                phone = str(row[col]) if pd.notna(row[col]) else ""
                properties["parent_phone_number"] = {"rich_text": [{"text": {"content": phone}}]}
            
            if "시작일" in df.columns or "start_date" in df.columns:
                col = "시작일" if "시작일" in df.columns else "start_date"
                date_value = row[col]
                if pd.notna(date_value):
                    try:
                        if isinstance(date_value, datetime):
                            date_obj = date_value
                        elif isinstance(date_value, str):
                            date_obj = pd.to_datetime(date_value)
                        else:
                            date_obj = pd.to_datetime(date_value)
                        properties["start_date"] = {"date": {"start": date_obj.strftime("%Y-%m-%d")}}
                    except:
                        pass
            
            if "학교명" in df.columns or "school_name" in df.columns:
                col = "학교명" if "학교명" in df.columns else "school_name"
                school_name = str(row[col]) if pd.notna(row[col]) else ""
                properties["school_name"] = {"rich_text": [{"text": {"content": school_name}}]}
            
            if "학년" in df.columns or "grade" in df.columns:
                col = "학년" if "학년" in df.columns else "grade"
                grade = row[col]
                if pd.notna(grade):
                    try:
                        # "3학년" 형식에서 숫자만 추출
                        if isinstance(grade, str):
                            grade_num = re.search(r'\d+', grade)
                            if grade_num:
                                properties["grade"] = {"number": int(grade_num.group())}
                        else:
                            properties["grade"] = {"number": int(grade)}
                    except:
                        pass
            
            # type -> type (특목, 본관 구분)
            if "type" in df.columns or "타입" in df.columns or "구분" in df.columns:
                col = None
                if "type" in df.columns:
                    col = "type"
                elif "타입" in df.columns:
                    col = "타입"
                elif "구분" in df.columns:
                    col = "구분"
                
                if col and pd.notna(row[col]):
                    type_value = str(row[col]).strip()
                    if type_value:
                        # rich_text 타입으로 처리 (특목, 본관)
                        properties["type"] = {"rich_text": [{"text": {"content": type_value}}]}
        
        elif table_type == "discharge":
            # discharge 테이블 속성 매핑 (지정된 컬럼명 사용)
            # 시작일 -> start_date
            if "시작일" in df.columns or "start_date" in df.columns:
                col = "시작일" if "시작일" in df.columns else "start_date"
                date_value = row[col]
                if pd.notna(date_value):
                    try:
                        if isinstance(date_value, datetime):
                            date_obj = date_value
                        elif isinstance(date_value, str):
                            date_obj = pd.to_datetime(date_value)
                        else:
                            date_obj = pd.to_datetime(date_value)
                        properties["start_date"] = {"date": {"start": date_obj.strftime("%Y-%m-%d")}}
                    except:
                        pass
            
            # 학생명 -> student_name
            if "학생명" in df.columns or "student_name" in df.columns:
                col = "학생명" if "학생명" in df.columns else "student_name"
                student_name = str(row[col]) if pd.notna(row[col]) else ""
                properties["student_name"] = {"title": [{"text": {"content": student_name}}]}
            
            # 부모HP -> parent_phone_number
            if "부모HP" in df.columns or "parent_phone_number" in df.columns:
                col = "부모HP" if "부모HP" in df.columns else "parent_phone_number"
                phone = str(row[col]) if pd.notna(row[col]) else ""
                properties["parent_phone_number"] = {"rich_text": [{"text": {"content": phone}}]}
            
            # 반명 -> class_name
            if "반명" in df.columns or "class_name" in df.columns:
                col = "반명" if "반명" in df.columns else "class_name"
                class_name = str(row[col]) if pd.notna(row[col]) else ""
                properties["class_name"] = {"rich_text": [{"text": {"content": class_name}}]}
            
            # 담당 -> teacher_name
            if "담당" in df.columns or "teacher_name" in df.columns:
                col = "담당" if "담당" in df.columns else "teacher_name"
                teacher_name = row[col]
                if pd.notna(teacher_name):
                    if isinstance(teacher_name, str):
                        # 쉼표로 구분된 경우 리스트로 변환
                        if "," in teacher_name:
                            teacher_list = [t.strip() for t in teacher_name.split(",")]
                            properties["teacher_name"] = {"multi_select": [{"name": str(t)} for t in teacher_list]}
                        else:
                            properties["teacher_name"] = {"rich_text": [{"text": {"content": str(teacher_name)}}]}
                    elif isinstance(teacher_name, list):
                        properties["teacher_name"] = {"multi_select": [{"name": str(t)} for t in teacher_name]}
            
            # 퇴원사유 -> discharging_reason
            if "퇴원사유" in df.columns or "discharging_reason" in df.columns:
                col = "퇴원사유" if "퇴원사유" in df.columns else "discharging_reason"
                reason = str(row[col]) if pd.notna(row[col]) else ""
                properties["discharging_reason"] = {"rich_text": [{"text": {"content": reason}}]}
            
            # 학교명 -> school_name
            if "학교명" in df.columns or "school_name" in df.columns:
                col = "학교명" if "학교명" in df.columns else "school_name"
                school_name = str(row[col]) if pd.notna(row[col]) else ""
                properties["school_name"] = {"rich_text": [{"text": {"content": school_name}}]}
            
            # 학년 -> grade
            if "학년" in df.columns or "grade" in df.columns:
                col = "학년" if "학년" in df.columns else "grade"
                grade = row[col]
                if pd.notna(grade):
                    try:
                        # "3학년" 형식에서 숫자만 추출
                        if isinstance(grade, str):
                            grade_num = re.search(r'\d+', grade)
                            if grade_num:
                                properties["grade"] = {"number": int(grade_num.group())}
                        else:
                            properties["grade"] = {"number": int(grade)}
                    except:
                        pass
            
            # 퇴원일자 -> discharge_date
            if "퇴원일자" in df.columns or "discharge_date" in df.columns:
                col = "퇴원일자" if "퇴원일자" in df.columns else "discharge_date"
                date_value = row[col]
                if pd.notna(date_value):
                    try:
                        if isinstance(date_value, datetime):
                            date_obj = date_value
                        elif isinstance(date_value, str):
                            date_obj = pd.to_datetime(date_value)
                        else:
                            date_obj = pd.to_datetime(date_value)
                        properties["discharge_date"] = {"date": {"start": date_obj.strftime("%Y-%m-%d")}}
                    except:
                        pass
            
            # type -> type (특목, 본관 구분)
            if "type" in df.columns or "타입" in df.columns or "구분" in df.columns:
                col = None
                if "type" in df.columns:
                    col = "type"
                elif "타입" in df.columns:
                    col = "타입"
                elif "구분" in df.columns:
                    col = "구분"
                
                if col and pd.notna(row[col]):
                    type_value = str(row[col]).strip()
                    if type_value:
                        # rich_text 타입으로 처리 (특목, 본관)
                        properties["type"] = {"rich_text": [{"text": {"content": type_value}}]}
        
        elif table_type == "student":
            # student 테이블 속성 매핑 (필요한 속성 추가)
            if "학생명" in df.columns or "student_name" in df.columns:
                col = "학생명" if "학생명" in df.columns else "student_name"
                student_name = str(row[col]) if pd.notna(row[col]) else ""
                properties["student_name"] = {"title": [{"text": {"content": student_name}}]}
            
            # student 테이블의 다른 속성들도 필요에 따라 추가
        
        return properties
    
    async def add_preprocessed_data_to_notion(self, df: pd.DataFrame, table_type: str) -> int:
        """전처리된 DataFrame을 Notion DB에 추가
        
        Args:
            df: 전처리된 DataFrame
            table_type: 테이블 타입 (class, discharge, student)
        
        Returns:
            추가된 페이지 수
        """
        if df.empty:
            logger.warning(f"⚠️ [{table_type}] 추가할 데이터가 없습니다.")
            return 0
        
        db_id = self.notion.db_map.get(table_type.lower())
        if not db_id:
            logger.error(f"❌ 테이블을 찾을 수 없음: {table_type}")
            return 0
        
        logger.info(f"📤 [{table_type}] Notion DB에 데이터 추가 시작: {len(df)}개 행")
        
        added_count = 0
        failed_count = 0
        
        for idx, row in df.iterrows():
            try:
                # DataFrame 행을 Notion 속성으로 변환
                properties = self._convert_dataframe_row_to_notion_properties(row, table_type, df)
                
                if not properties:
                    logger.warning(f"⚠️ [{table_type}] 행 {idx}: 변환된 속성이 없어 건너뜁니다.")
                    continue
                
                # Notion에 추가
                await self.notion.client.pages.create(
                    parent={"database_id": db_id},
                    properties=properties
                )
                
                added_count += 1
                if added_count % 10 == 0:
                    logger.info(f"📝 [{table_type}] 진행 중: {added_count}/{len(df)}개 추가됨")
                
                # API 제한 고려 (초당 3회)
                await asyncio.sleep(0.35)
                
            except Exception as e:
                failed_count += 1
                logger.error(f"❌ [{table_type}] 행 {idx} 추가 실패: {e}")
        
        logger.info(f"✅ [{table_type}] Notion DB 추가 완료: 성공 {added_count}개, 실패 {failed_count}개")
        
        return added_count


    


####

class EnhancedDischargeReportGenerator:
    """차트 포함 월별 입퇴소 현황 생성기"""
    
    def __init__(self, notion_manager):
        self.notion = notion_manager
    
    async def generate_monthly_report(self, query_results,
                                      teacher_name: str,
                                      year: Optional[int] = None,
                                      month: Optional[int] = None
                                      ) -> Dict:
        """
        월별 입퇴소 현황 + 12개월 추이 데이터 생성
        
        Returns:
            {
                "current_month": {...},     # 해당 월 상세
                "yearly_trend": {...},       # 12개월 추이
                "detailed_list": [...]       # 학생별 상세 명단
            }
        """
        # 년월이 제공되지 않으면 현재 날짜 사용
        if year is None:
            year = datetime.now().year
        if month is None:
            month = datetime.now().month
        
        logger.info(f"📊 {teacher_name} {year}년 {month}월 입퇴소 현황 생성")
        
        # 1. 해당 월 데이터
        current_data = await self._get_current_month_data(
            query_results, year, month
        )
        
        # 2. 12개월 추이 데이터 (과거 11개월 + 현재월)
        yearly_trend = await self._get_yearly_trend(
            query_results, teacher_name, year, month
        )
        
        # 3. 학생별 상세 명단 (입소일, 퇴소일 포함)
        detailed_list = await self._get_detailed_student_list(
            query_results
        )
        
        return {
            "teacher_name": teacher_name,
            "year": year,
            "month": month,
            "current_month": current_data,
            "yearly_trend": yearly_trend,
            "detailed_list": detailed_list
        }

    async def _get_current_month_data(self, query_results,
                                     year: int, 
                                     month: int) -> Dict:
        """해당 월 입퇴소 데이터"""

        # 입소 데이터 (class 테이블)
        enrollments = await self.year_month_enrollment(
            query_results, year, month
        )
        
        # 퇴소 데이터 (discharge 테이블)
        discharges = await self.year_month_discharge(
            query_results, year, month
        )
        
        return {
            "enrollments": len(enrollments),
            "discharges": len(discharges),
            "net_change": len(enrollments) - len(discharges),
            "enrollment_list": enrollments,
            "discharge_list": discharges
        }
       
    async def _get_yearly_trend(self, query_results,
                                teacher_name: str, 
                                year: int, 
                                month: int) -> Dict:
        """가용 데이터 개월 수에 맞춘 월별 추이 데이터"""
        trend_data: List[Dict[str, Any]] = []

        # 1) 데이터에 존재하는 월 수집 (class: start_date, discharge: discharge_date)
        months_set: set = set()

        def _add_month(val):
            if not val:
                return
            if isinstance(val, list):
                for v in val:
                    _add_month(v)
                return
            try:
                d = datetime.fromisoformat(str(val).split("T")[0])
                months_set.add((d.year, d.month))
            except Exception:
                return

        if isinstance(query_results, dict):
            normalized = {k.lower(): v for k, v in query_results.items()}
            for item in normalized.get("class", []):
                v = item.get("start_date") or item.get("start") or item.get("입소일") or item.get("startDate")
                _add_month(v)
            for item in normalized.get("discharge", []):
                v = item.get("discharge_date") or item.get("discharge") or item.get("퇴소일") or item.get("dischargeDate")
                _add_month(v)

        # 2) 사용할 월 목록 결정
        if months_set:
            month_targets = sorted(months_set)  # (year, month) 오름차순
        else:
            # 데이터가 없으면 기존처럼 최근 12개월을 사용
            month_targets = []
            for i in range(11, -1, -1):
                target_date = datetime(year, month, 1) - timedelta(days=i*30)
                month_targets.append((target_date.year, target_date.month))

        # 3) 월별 입소/퇴소 집계
        for target_year, target_month in month_targets:
            start_date, end_date = self._get_month_range(target_year, target_month)

            enrollments = await self.year_month_enrollment(
                query_results, target_year, target_month
            )
            discharges = await self.year_month_discharge(
                query_results, target_year, target_month
            )

            logger.debug(f"[Trend] {target_year}-{target_month:02d} enrollments={len(enrollments)} discharges={len(discharges)}")

            trend_data.append({
                "year": target_year,
                "month": target_month,
                "month_label": f"{target_year}년 {target_month}월",
                "enrollments": len(enrollments),
                "discharges": len(discharges),
                "net_change": len(enrollments) - len(discharges)
            })

        # 4) 차트 타입 결정: 데이터가 한 개월뿐이면 막대형
        chart_type = "bar" if len(trend_data) == 1 else "line"

        return {
            "monthly_data": trend_data,
            "chart_type": chart_type
        }
    
    async def _get_detailed_student_list(self, query_results
                                        ) -> List[Dict]:
        """학생별 상세 명단 (입소일, 퇴소일 포함)"""       
        detailed_list = []
        
        # 1. 입소 학생 (class 테이블 - 퇴소일 없음)
        enrollments = query_results.get("class", [])

        for student in enrollments:
            # enrollments now use internal English keys; map to output Korean keys
            start_val = student.get("start_date")
            detailed_list.append({
                "학생명": student.get("student_name"),
                "학년": f"{student.get('grade')}학년" if isinstance(student.get('grade'), int) else student.get('grade'),
                "반": student.get("class_name"),
                "입소일자": start_val,
                "퇴소일자": None,
                "재원상태": "재원중",
                "재원기간": self._calculate_days_from(start_val),
                "퇴원사유": None,
                "학부모전화": student.get("parent_phone_number")
            })
        
        # 2. 퇴소 학생 (discharge 테이블 - 입소일 + 퇴소일 있음)
        discharges = query_results.get("discharge", [])

        for student in discharges:
            start_val = student.get("start_date")
            end_val = student.get("discharge_date")
            detailed_list.append({
                "학생명": student.get("student_name"),
                "학년": f"{student.get('grade')}학년" if isinstance(student.get('grade'), int) else student.get('grade'),
                "반": ", ".join(student.get("class_name", [])) if isinstance(student.get("class_name"), list) else student.get("class_name", ""),
                "입소일자": start_val,
                "퇴소일자": end_val,
                "재원상태": "퇴원",
                "재원기간": self._calculate_duration(
                    start_val, 
                    end_val
                ),
                "퇴원사유": student.get("discharging_reason"),
                "학부모전화": student.get("parent_phone_number")
        })
        
        # 퇴소일자 기준 정렬 (퇴소일자가 없으면 입소일자 사용)
        detailed_list.sort(
            key=lambda x: x.get("퇴소일자") or x.get("입소일자") or "9999-99-99"
        )
        
        return detailed_list
    
   
            
    async def year_month_enrollment(self, query_results: Dict, year: int, month: int) -> List[Dict]:
        """Compatibility wrapper for requested name `year_month_enrollment`."""
        # Normalize keys and pull class list
        if isinstance(query_results, dict):
            normalized = {k.lower(): v for k, v in query_results.items()}
        else:
            return []

        class_list = normalized.get("class", [])
        start_range, end_range = self._get_month_range(year, month)

        filtered: List[Dict] = []
        for item in class_list:
            # possible date keys
            val = item.get("start_date") or item.get("start") or item.get("입소일") or item.get("startDate")
            # support nested Notion-like dicts
            if isinstance(val, dict):
                val = val.get("date") or val.get("start") or val.get("start_date")
                if isinstance(val, dict):
                    val = val.get("start")
            if not val:
                continue
            try:
                d = datetime.fromisoformat(str(val).split("T")[0])
            except Exception:
                continue
            if start_range.date() <= d.date() <= end_range.date():
                filtered.append(item)
        return filtered

    async def year_month_discharge(self, query_results: Dict, year: int, month: int) -> List[Dict]:
        """Compatibility wrapper for requested name `year_month_discharge`."""
        # Normalize keys and pull discharge list
        if isinstance(query_results, dict):
            normalized = {k.lower(): v for k, v in query_results.items()}
        else:
            return []

        discharge_list = normalized.get("discharge", [])
        start_range, end_range = self._get_month_range(year, month)

        filtered: List[Dict] = []
        for item in discharge_list:
            val = item.get("discharge_date") or item.get("discharge") or item.get("퇴소일") or item.get("dischargeDate")
            if isinstance(val, dict):
                val = val.get("date") or val.get("start") or val.get("discharge_date")
                if isinstance(val, dict):
                    val = val.get("start")
            if not val:
                continue
            try:
                d = datetime.fromisoformat(str(val).split("T")[0])
            except Exception:
                continue
            if start_range.date() <= d.date() <= end_range.date():
                filtered.append(item)
        return filtered

    
    def _get_month_range(self, year: int, month: int) -> Tuple[datetime, datetime]:
        """해당 월의 시작일/종료일"""
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1) - timedelta(days=1)
        return start_date, end_date
    
    def _calculate_duration(self, start_str: str, end_str: str) -> str:
        """재원 기간 계산"""
        try:
            start = datetime.fromisoformat(start_str.split("T")[0])
            end = datetime.fromisoformat(end_str.split("T")[0])
            days = (end - start).days
            
            if days < 30:
                return f"{days}일"
            else:
                months = days // 30
                remaining = days % 30
                if remaining > 0:
                    return f"{months}개월 {remaining}일"
                return f"{months}개월"
        except:
            return "-"
        
    def _calculate_days_from(self, start_str: str) -> str:
        """입소일부터 현재까지"""
        try:
            start = datetime.fromisoformat(start_str.split("T")[0])
            days = (datetime.now() - start).days
            
            if days < 30:
                return f"{days}일"
            else:
                months = days // 30
                remaining = days % 30
                return f"{months}개월 {remaining}일"
        except:
            return "-"
        
    
    def create_excel_with_chart(self, report_data: Dict, 
                                filename: str) -> Path:
        """차트 포함 Excel 생성"""
        logger.info("📊 차트 포함 Excel 생성 중...")
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # ===== 시트 1: 월별 추이 (차트 포함) =====
        ws_trend = wb.create_sheet("월별 추이")
        self._create_trend_sheet_with_chart(ws_trend, report_data)

        # ===== 시트 2: 월별 요약 =====
        ws_summary = wb.create_sheet("월별 요약")
        self._create_summary_sheet(ws_summary, report_data)
        
        # ===== 시트 3: 과목별 입퇴소 추이 시트 + 차트 =====
        ws_class_trend = wb.create_sheet("과목별 입퇴소 추이")
        self._create_class_trend_sheet_with_chart(ws_class_trend, report_data)

        # ===== 시트 4: 과목별 퇴소 사유 순위 요약 시트 =====
        ws_class_summary = wb.create_sheet("과목별 퇴소 사유 순위 요약")
        self._create_class_summary_sheet(ws_class_summary, report_data)
        
        # ===== 시트 5: 학생별 상세 명단 =====
        ws_detail = wb.create_sheet("학생 상세")
        self._create_detail_sheet(ws_detail, report_data)

        
        
        # 저장
        output_path = Path("temp") / f"{filename}.xlsx"
        wb.save(output_path)
        
        logger.info(f"✅ 차트 포함 Excel 생성 완료: {output_path.name}")
        return output_path
    
    def _create_trend_sheet_with_chart(self, ws, report_data: Dict):
        """월별 추이 시트 + 차트 (데이터 개월수 기반)"""
        trend_info = report_data.get("yearly_trend", {})
        trend_data = trend_info.get("monthly_data", [])
        chart_type = trend_info.get("chart_type", "line")
        month_count = len(trend_data)
        
        # 제목
        ws.merge_cells('A1:G1')
        title = ws['A1']
        title.value = f"📈 {report_data['teacher_name']} - {month_count}개월 입퇴소 추이"
        title.font = Font(size=16, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # 헤더
        headers = ["월", "입소", "퇴소", "순증감"]
        header_row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # 데이터
        for row_idx, data in enumerate(trend_data, header_row + 1):
            ws.cell(row=row_idx, column=1).value = f"{data['year']}.{data['month']:02d}"
            ws.cell(row=row_idx, column=2).value = data['enrollments']
            ws.cell(row=row_idx, column=3).value = data['discharges']
            ws.cell(row=row_idx, column=4).value = data['net_change']
            
            # 순증감 색상
            net_cell = ws.cell(row=row_idx, column=4)
            if data['net_change'] > 0:
                net_cell.font = Font(color="00B050", bold=True)
            elif data['net_change'] < 0:
                net_cell.font = Font(color="FF0000", bold=True)
            
            # 스트라이프
            if row_idx % 2 == 0:
                for col in range(1, 5):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                    )
        
        # 열 너비
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        
        # ===== 차트 생성 =====
        # 데이터 범위 (월, 입소, 퇴소)
        data_ref = Reference(ws, min_col=2, min_row=header_row, 
                             max_row=header_row + len(trend_data), max_col=3)
        cats_ref = Reference(ws, min_col=1, min_row=header_row + 1, 
                             max_row=header_row + len(trend_data))

        if chart_type == "bar":
            # 데이터가 한 개월뿐이면 막대형으로 표현
            bar_chart_main = BarChart()
            bar_chart_main.type = "col"
            bar_chart_main.title = "월별 입퇴소"
            bar_chart_main.y_axis.title = "인원 (명)"
            bar_chart_main.x_axis.title = "월"
            bar_chart_main.height = 12
            bar_chart_main.width = 24
            bar_chart_main.add_data(data_ref, titles_from_data=True)
            bar_chart_main.set_categories(cats_ref)
            bar_chart_main.dataLabels = DataLabelList()
            bar_chart_main.dataLabels.showVal = True
            ws.add_chart(bar_chart_main, f"F3")
        else:
            # 기본: 꺾은선 차트 (입소/퇴소)
            line_chart = LineChart()
            line_chart.title = "월별 입퇴소 추이"
            line_chart.style = 13
            line_chart.y_axis.title = "인원 (명)"
            line_chart.x_axis.title = "월"
            line_chart.height = 12
            line_chart.width = 24
            
            line_chart.add_data(data_ref, titles_from_data=True)
            line_chart.set_categories(cats_ref)
            try:
                line_chart.series[0].graphicalProperties.line.solidFill = "00B050"
                line_chart.series[1].graphicalProperties.line.solidFill = "FF0000"
            except Exception:
                pass
            
            line_chart.dataLabels = DataLabelList()
            line_chart.dataLabels.showVal = True
            ws.add_chart(line_chart, f"F3")
        
        # 막대 차트 (순증감)
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.title = "월별 순증감"
        bar_chart.y_axis.title = "순증감 (명)"
        bar_chart.x_axis.title = "월"
        bar_chart.height = 12
        bar_chart.width = 24
        
        # 순증감 데이터
        data = Reference(ws, min_col=4, min_row=header_row, 
                        max_row=header_row + len(trend_data))
        cats = Reference(ws, min_col=1, min_row=header_row + 1, 
                        max_row=header_row + len(trend_data))
        
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(cats)
        
        # 차트 삽입
        ws.add_chart(bar_chart, f"F23")
    
    def _create_class_trend_sheet_with_chart(self, ws, report_data: Dict):
        """과목별 입퇴소 추이 시트 + 차트"""
        detailed_list = report_data.get("detailed_list", [])
        
        # 과목별 입소/퇴소 집계
        subject_enrollments = Counter()
        subject_discharges = Counter()
        
        for student in detailed_list:
            subject = student.get("반", "기타")
            # 반이 리스트인 경우 처리
            if isinstance(subject, list):
                subjects = subject
            elif isinstance(subject, str) and "," in subject:
                subjects = [s.strip() for s in subject.split(",")]
            else:
                subjects = [subject] if subject else ["기타"]
            
            for subj in subjects:
                if student.get("재원상태") == "재원중":
                    subject_enrollments[subj] += 1
                elif student.get("재원상태") == "퇴원":
                    subject_discharges[subj] += 1
        
        # 모든 과목 수집 및 정렬
        all_subjects = set(subject_enrollments.keys()) | set(subject_discharges.keys())
        subject_data = []
        for subject in sorted(all_subjects):
            enrollments = subject_enrollments.get(subject, 0)
            discharges = subject_discharges.get(subject, 0)
            net_change = enrollments - discharges
            subject_data.append({
                "subject": subject,
                "enrollments": enrollments,
                "discharges": discharges,
                "net_change": net_change
            })
        
        # 제목
        ws.merge_cells('A1:G1')
        title = ws['A1']
        title.value = f"📚 {report_data['teacher_name']} - 과목별 입퇴소 추이"
        title.font = Font(size=16, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # 헤더
        headers = ["과목", "입소", "퇴소", "순증감"]
        header_row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # 데이터
        for row_idx, data in enumerate(subject_data, header_row + 1):
            ws.cell(row=row_idx, column=1).value = data['subject']
            ws.cell(row=row_idx, column=2).value = data['enrollments']
            ws.cell(row=row_idx, column=3).value = data['discharges']
            ws.cell(row=row_idx, column=4).value = data['net_change']
            
            # 순증감 색상
            net_cell = ws.cell(row=row_idx, column=4)
            if data['net_change'] > 0:
                net_cell.font = Font(color="00B050", bold=True)
            elif data['net_change'] < 0:
                net_cell.font = Font(color="FF0000", bold=True)
            
            # 스트라이프
            if row_idx % 2 == 0:
                for col in range(1, 5):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                    )
        
        # 열 너비
        ws.column_dimensions['A'].width = 20  # 과목명은 더 넓게
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        
        # ===== 차트 생성 =====
        # 막대 차트 (입소/퇴소)
        bar_chart_enroll = BarChart()
        bar_chart_enroll.type = "col"
        bar_chart_enroll.title = "과목별 입퇴소 추이"
        bar_chart_enroll.style = 13
        bar_chart_enroll.y_axis.title = "인원 (명)"
        bar_chart_enroll.x_axis.title = "과목"
        bar_chart_enroll.height = 12
        bar_chart_enroll.width = 24
        
        # 데이터 범위
        data = Reference(ws, min_col=2, min_row=header_row, 
                        max_row=header_row + len(subject_data), max_col=3)
        cats = Reference(ws, min_col=1, min_row=header_row + 1, 
                        max_row=header_row + len(subject_data))
        
        bar_chart_enroll.add_data(data, titles_from_data=True)
        bar_chart_enroll.set_categories(cats)
        # 색상 지정: series[0]=입소(초록), series[1]=퇴소(붉은)
        try:
            bar_chart_enroll.series[0].graphicalProperties.solidFill = "00B050"
            bar_chart_enroll.series[1].graphicalProperties.solidFill = "FF0000"
        except Exception:
            pass
        
        # 데이터 레이블 표시
        bar_chart_enroll.dataLabels = DataLabelList()
        bar_chart_enroll.dataLabels.showVal = True
        
        # 차트 삽입 위치
        ws.add_chart(bar_chart_enroll, f"F3")
        
        # 막대 차트 (순증감)
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.title = "과목별 순증감"
        bar_chart.y_axis.title = "순증감 (명)"
        bar_chart.x_axis.title = "과목"
        bar_chart.height = 12
        bar_chart.width = 24
        
        # 순증감 데이터
        data = Reference(ws, min_col=4, min_row=header_row, 
                        max_row=header_row + len(subject_data))
        cats = Reference(ws, min_col=1, min_row=header_row + 1, 
                        max_row=header_row + len(subject_data))
        
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(cats)
        
        # 차트 삽입
        ws.add_chart(bar_chart, f"F23")
    
    def _create_summary_sheet(self, ws, report_data: Dict):
        """모든 월 요약 시트"""
        
        # 제목
        ws.merge_cells('A1:G1')
        title = ws['A1']
        title.value = f"📊 월별 퇴소 사유 순위 요약"
        title.font = Font(size=16, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # detailed_list에서 퇴원한 학생들만 필터링
        detailed_list = report_data.get("detailed_list", [])
        discharged_students = [
            student for student in detailed_list 
            if student.get("재원상태") == "퇴원" and student.get("퇴소일자")
        ]
        
        if not discharged_students:
            ws.cell(row=3, column=1).value = "퇴소 데이터가 없습니다."
            return
        
        # 퇴소일자를 기준으로 월별로 그룹화
        monthly_discharges = {}
        for student in discharged_students:
            discharge_date_str = student.get("퇴소일자")
            if not discharge_date_str:
                continue
            
            try:
                # 날짜 문자열 파싱 (ISO 형식 또는 다른 형식 지원)
                if "T" in discharge_date_str:
                    discharge_date = datetime.fromisoformat(discharge_date_str.split("T")[0])
                else:
                    discharge_date = datetime.fromisoformat(discharge_date_str)
                
                year_month = (discharge_date.year, discharge_date.month)
                if year_month not in monthly_discharges:
                    monthly_discharges[year_month] = []
                
                monthly_discharges[year_month].append(student)
            except Exception:
                continue
        
        # 월별로 정렬 (년도, 월 순서)
        sorted_months = sorted(monthly_discharges.keys())
        
        # 각 월별 데이터를 먼저 준비
        monthly_data = []
        for year, month in sorted_months:
            students = monthly_discharges[(year, month)]
            
            # 퇴소 사유 집계
            reasons = []
            for student in students:
                r = student.get('퇴원사유') or student.get('퇴소사유') or student.get('discharging_reason')
                if not r:
                    r = '기타'
                if isinstance(r, str):
                    r = r.strip() or '기타'
                else:
                    r = str(r)
                reasons.append(r)
            
            counts = Counter(reasons)
            sorted_reasons = counts.most_common()
            
            monthly_data.append({
                'year': year,
                'month': month,
                'reasons': sorted_reasons
            })
        
        # 3행 4열 그리드로 배치
        # 각 월별 박스는 4열 너비 (A~D, E~H, I~L, M~P)
        # 각 행의 시작 열: A(1), E(5), I(9), M(13)
        cols_per_month = 4
        start_cols = [1, 5, 9, 13]  # A, E, I, M
        
        # 각 월별 박스의 최대 높이 계산 (제목 1행 + 헤더 1행 + 데이터 최대 10행)
        max_rows_per_month = 12
        
        # 3행 4열 그리드로 배치
        for month_idx, month_info in enumerate(monthly_data):
            year = month_info['year']
            month = month_info['month']
            sorted_reasons = month_info['reasons']
            
            # 그리드 위치 계산 (3행 4열)
            grid_row = month_idx // 4  # 0, 1, 2
            grid_col = month_idx % 4   # 0, 1, 2, 3
            
            # 실제 Excel 행/열 계산
            start_col = start_cols[grid_col]
            start_row = 3 + (grid_row * max_rows_per_month)
            
            # 월별 섹션 제목
            end_col = start_col + cols_per_month - 1
            ws.merge_cells(f'{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{start_row}')
            sec_title = ws[f'{get_column_letter(start_col)}{start_row}']
            sec_title.value = f"📅 {year}년 {month}월"
            sec_title.font = Font(size=11, bold=True, color="FFFFFF")
            sec_title.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            sec_title.alignment = Alignment(horizontal="center", vertical="center")
            
            current_row = start_row + 1
            
            if sorted_reasons:
                # 테이블 헤더
                headers = ["순위", "사유", "건수"]
                for col_offset, header in enumerate(headers):
                    col = start_col + col_offset
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = header
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True, size=9)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                current_row += 1
                
                # 데이터 행 (최대 10개까지만 표시)
                for idx, (reason, cnt) in enumerate(sorted_reasons[:10], 1):
                    ws.cell(row=current_row, column=start_col).value = idx
                    ws.cell(row=current_row, column=start_col + 1).value = reason[:15] if len(reason) > 15 else reason  # 사유는 최대 15자
                    ws.cell(row=current_row, column=start_col + 2).value = f"{cnt}건"
                    
                    # 스트라이프 스타일
                    if current_row % 2 == 0:
                        for c in range(start_col, start_col + 3):
                            ws.cell(row=current_row, column=c).fill = PatternFill(
                                start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                            )
                    
                    # 작은 폰트
                    for c in range(start_col, start_col + 3):
                        ws.cell(row=current_row, column=c).font = Font(size=9)
                        ws.cell(row=current_row, column=c).alignment = Alignment(horizontal="center", vertical="center")
                    
                    current_row += 1
            else:
                ws.cell(row=current_row, column=start_col).value = "데이터 없음"
                ws.cell(row=current_row, column=start_col).font = Font(size=9)
                ws.cell(row=current_row, column=start_col).alignment = Alignment(horizontal="center", vertical="center")
        
        # 열 너비 조정 (각 월별 박스의 열)
        for col_letter in ['A', 'E', 'I', 'M']:
            ws.column_dimensions[col_letter].width = 6  # 순위
        for col_letter in ['B', 'F', 'J', 'N']:
            ws.column_dimensions[col_letter].width = 18  # 사유
        for col_letter in ['C', 'G', 'K', 'O']:
            ws.column_dimensions[col_letter].width = 8   # 건수
        # 빈 열 (간격용)
        for col_letter in ['D', 'H', 'L', 'P']:
            ws.column_dimensions[col_letter].width = 2
        
        # ===== 전체 월 퇴소사유 요약 (오른쪽) =====
        # 전체 퇴소사유 집계
        all_reasons = []
        for student in discharged_students:
            r = student.get('퇴원사유') or student.get('퇴소사유') or student.get('discharging_reason')
            if not r:
                r = '기타'
            if isinstance(r, str):
                r = r.strip() or '기타'
            else:
                r = str(r)
            all_reasons.append(r)
        
        all_counts = Counter(all_reasons)
        all_sorted_reasons = all_counts.most_common()
        
        # 전체 요약 섹션 시작 위치 (R열부터, 한 칸 더 띄움)
        summary_start_col = 18  # R열 (Q열에서 한 칸 오른쪽)
        summary_start_row = 3
        
        # 전체 요약 제목 (2행 높이로 확장)
        ws.merge_cells(f'{get_column_letter(summary_start_col)}{summary_start_row}:{get_column_letter(summary_start_col + 2)}{summary_start_row + 1}')
        summary_title = ws[f'{get_column_letter(summary_start_col)}{summary_start_row}']
        summary_title.value = "📊 전체 월 퇴소사유 요약"
        summary_title.font = Font(size=16, bold=True, color="FFFFFF")
        summary_title.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")  # 주황색 계열로 변경
        summary_title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[summary_start_row].height = 40  # 제목 행 높이 증가
        ws.row_dimensions[summary_start_row + 1].height = 40  # 제목 행 높이 증가
        
        current_summary_row = summary_start_row + 2
        
        if all_sorted_reasons:
            # 테이블 헤더
            headers = ["순위", "사유", "건수"]
            for col_offset, header in enumerate(headers):
                col = summary_start_col + col_offset
                cell = ws.cell(row=current_summary_row, column=col)
                cell.value = header
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            current_summary_row += 1
            
            # 데이터 행
            for idx, (reason, cnt) in enumerate(all_sorted_reasons, 1):
                ws.cell(row=current_summary_row, column=summary_start_col).value = idx
                ws.cell(row=current_summary_row, column=summary_start_col + 1).value = reason
                ws.cell(row=current_summary_row, column=summary_start_col + 2).value = f"{cnt}건"
                
                # 스트라이프 스타일
                if current_summary_row % 2 == 0:
                    for c in range(summary_start_col, summary_start_col + 3):
                        ws.cell(row=current_summary_row, column=c).fill = PatternFill(
                            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                        )
                
                # 폰트 및 정렬
                for c in range(summary_start_col, summary_start_col + 3):
                    ws.cell(row=current_summary_row, column=c).font = Font(size=10)
                    ws.cell(row=current_summary_row, column=c).alignment = Alignment(horizontal="center", vertical="center")
                
                current_summary_row += 1
        else:
            ws.cell(row=current_summary_row, column=summary_start_col).value = "데이터 없음"
            ws.cell(row=current_summary_row, column=summary_start_col).font = Font(size=10)
            ws.cell(row=current_summary_row, column=summary_start_col).alignment = Alignment(horizontal="center", vertical="center")
        
        # 전체 요약 열 너비 조정 (R, S, T열)
        ws.column_dimensions['R'].width = 6   # 순위
        ws.column_dimensions['S'].width = 25  # 사유
        ws.column_dimensions['T'].width = 10  # 건수
    
    def _create_class_summary_sheet(self, ws, report_data: Dict):
        """과목별 퇴소 사유 순위 요약 시트"""
        
        # 제목
        ws.merge_cells('A1:G1')
        title = ws['A1']
        title.value = f"📊 과목별 퇴소 사유 순위 요약"
        title.font = Font(size=16, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # detailed_list에서 퇴원한 학생들만 필터링
        detailed_list = report_data.get("detailed_list", [])
        discharged_students = [
            student for student in detailed_list 
            if student.get("재원상태") == "퇴원" and student.get("퇴소일자")
        ]
        
        if not discharged_students:
            ws.cell(row=3, column=1).value = "퇴소 데이터가 없습니다."
            return
        
        # 과목별로 그룹화
        subject_discharges = {}
        for student in discharged_students:
            subject = student.get("반", "기타")
            # 반이 리스트인 경우 처리
            if isinstance(subject, list):
                subjects = subject
            elif isinstance(subject, str) and "," in subject:
                subjects = [s.strip() for s in subject.split(",")]
            else:
                subjects = [subject] if subject else ["기타"]
            
            for subj in subjects:
                if subj not in subject_discharges:
                    subject_discharges[subj] = []
                subject_discharges[subj].append(student)
        
        # 과목별로 정렬
        sorted_subjects = sorted(subject_discharges.keys())
        
        # 각 과목별 데이터를 먼저 준비
        subject_data = []
        for subject in sorted_subjects:
            students = subject_discharges[subject]
            
            # 퇴소 사유 집계
            reasons = []
            for student in students:
                r = student.get('퇴원사유') or student.get('퇴소사유') or student.get('discharging_reason')
                if not r:
                    r = '기타'
                if isinstance(r, str):
                    r = r.strip() or '기타'
                else:
                    r = str(r)
                reasons.append(r)
            
            counts = Counter(reasons)
            sorted_reasons = counts.most_common()
            
            subject_data.append({
                'subject': subject,
                'reasons': sorted_reasons
            })
        
        # 3행 4열 그리드로 배치
        # 각 과목별 박스는 4열 너비 (A~D, E~H, I~L, M~P)
        # 각 행의 시작 열: A(1), E(5), I(9), M(13)
        cols_per_subject = 4
        start_cols = [1, 5, 9, 13]  # A, E, I, M
        
        # 각 과목별 박스의 최대 높이 계산 (제목 1행 + 헤더 1행 + 데이터 최대 10행)
        max_rows_per_subject = 12
        
        # 3행 4열 그리드로 배치
        for subject_idx, subject_info in enumerate(subject_data):
            subject = subject_info['subject']
            sorted_reasons = subject_info['reasons']
            
            # 그리드 위치 계산 (3행 4열)
            grid_row = subject_idx // 4  # 0, 1, 2
            grid_col = subject_idx % 4   # 0, 1, 2, 3
            
            # 실제 Excel 행/열 계산
            start_col = start_cols[grid_col]
            start_row = 3 + (grid_row * max_rows_per_subject)
            
            # 과목별 섹션 제목
            end_col = start_col + cols_per_subject - 1
            ws.merge_cells(f'{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{start_row}')
            sec_title = ws[f'{get_column_letter(start_col)}{start_row}']
            sec_title.value = f"📚 {subject}"
            sec_title.font = Font(size=11, bold=True, color="FFFFFF")
            sec_title.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
            sec_title.alignment = Alignment(horizontal="center", vertical="center")
            
            current_row = start_row + 1
            
            if sorted_reasons:
                # 테이블 헤더
                headers = ["순위", "사유", "건수"]
                for col_offset, header in enumerate(headers):
                    col = start_col + col_offset
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = header
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True, size=9)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                current_row += 1
                
                # 데이터 행 (최대 10개까지만 표시)
                for idx, (reason, cnt) in enumerate(sorted_reasons[:10], 1):
                    ws.cell(row=current_row, column=start_col).value = idx
                    ws.cell(row=current_row, column=start_col + 1).value = reason[:15] if len(reason) > 15 else reason  # 사유는 최대 15자
                    ws.cell(row=current_row, column=start_col + 2).value = f"{cnt}건"
                    
                    # 스트라이프 스타일
                    if current_row % 2 == 0:
                        for c in range(start_col, start_col + 3):
                            ws.cell(row=current_row, column=c).fill = PatternFill(
                                start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                            )
                    
                    # 작은 폰트
                    for c in range(start_col, start_col + 3):
                        ws.cell(row=current_row, column=c).font = Font(size=9)
                        ws.cell(row=current_row, column=c).alignment = Alignment(horizontal="center", vertical="center")
                    
                    current_row += 1
            else:
                ws.cell(row=current_row, column=start_col).value = "데이터 없음"
                ws.cell(row=current_row, column=start_col).font = Font(size=9)
                ws.cell(row=current_row, column=start_col).alignment = Alignment(horizontal="center", vertical="center")
        
        # 열 너비 조정 (각 과목별 박스의 열)
        for col_letter in ['A', 'E', 'I', 'M']:
            ws.column_dimensions[col_letter].width = 6  # 순위
        for col_letter in ['B', 'F', 'J', 'N']:
            ws.column_dimensions[col_letter].width = 18  # 사유
        for col_letter in ['C', 'G', 'K', 'O']:
            ws.column_dimensions[col_letter].width = 8   # 건수
        # 빈 열 (간격용)
        for col_letter in ['D', 'H', 'L', 'P']:
            ws.column_dimensions[col_letter].width = 2
        
        # ===== 전체 과목 퇴소사유 요약 (오른쪽) =====
        # 전체 퇴소사유 집계
        all_reasons = []
        for student in discharged_students:
            r = student.get('퇴원사유') or student.get('퇴소사유') or student.get('discharging_reason')
            if not r:
                r = '기타'
            if isinstance(r, str):
                r = r.strip() or '기타'
            else:
                r = str(r)
            all_reasons.append(r)
        
        all_counts = Counter(all_reasons)
        all_sorted_reasons = all_counts.most_common()
        
        # 전체 요약 섹션 시작 위치 (R열부터, 한 칸 더 띄움)
        summary_start_col = 18  # R열 (Q열에서 한 칸 오른쪽)
        summary_start_row = 3
        
        # 전체 요약 제목 (2행 높이로 확장)
        ws.merge_cells(f'{get_column_letter(summary_start_col)}{summary_start_row}:{get_column_letter(summary_start_col + 2)}{summary_start_row + 1}')
        summary_title = ws[f'{get_column_letter(summary_start_col)}{summary_start_row}']
        summary_title.value = "📊 전체 과목 퇴소사유 요약"
        summary_title.font = Font(size=16, bold=True, color="FFFFFF")
        summary_title.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")  # 주황색 계열로 변경
        summary_title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[summary_start_row].height = 40  # 제목 행 높이 증가
        ws.row_dimensions[summary_start_row + 1].height = 40  # 제목 행 높이 증가
        
        current_summary_row = summary_start_row + 2
        
        if all_sorted_reasons:
            # 테이블 헤더
            headers = ["순위", "사유", "건수"]
            for col_offset, header in enumerate(headers):
                col = summary_start_col + col_offset
                cell = ws.cell(row=current_summary_row, column=col)
                cell.value = header
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            current_summary_row += 1
            
            # 데이터 행
            for idx, (reason, cnt) in enumerate(all_sorted_reasons, 1):
                ws.cell(row=current_summary_row, column=summary_start_col).value = idx
                ws.cell(row=current_summary_row, column=summary_start_col + 1).value = reason
                ws.cell(row=current_summary_row, column=summary_start_col + 2).value = f"{cnt}건"
                
                # 스트라이프 스타일
                if current_summary_row % 2 == 0:
                    for c in range(summary_start_col, summary_start_col + 3):
                        ws.cell(row=current_summary_row, column=c).fill = PatternFill(
                            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                        )
                
                # 폰트 및 정렬
                for c in range(summary_start_col, summary_start_col + 3):
                    ws.cell(row=current_summary_row, column=c).font = Font(size=10)
                    ws.cell(row=current_summary_row, column=c).alignment = Alignment(horizontal="center", vertical="center")
                
                current_summary_row += 1
        else:
            ws.cell(row=current_summary_row, column=summary_start_col).value = "데이터 없음"
            ws.cell(row=current_summary_row, column=summary_start_col).font = Font(size=10)
            ws.cell(row=current_summary_row, column=summary_start_col).alignment = Alignment(horizontal="center", vertical="center")
        
        # 전체 요약 열 너비 조정 (R, S, T열)
        ws.column_dimensions['R'].width = 6   # 순위
        ws.column_dimensions['S'].width = 25  # 사유
        ws.column_dimensions['T'].width = 10  # 건수
    
    def _parse_duration_to_days(self, duration_str: str) -> int:
        """재원기간 문자열을 일수로 변환"""
        if not duration_str or duration_str == "-":
            return 0
        try:
            days = 0
            # "X개월 Y일" 형식 파싱
            if "개월" in duration_str:
                months = int(re.search(r'(\d+)개월', duration_str).group(1))
                days += months * 30
            if "일" in duration_str:
                day_part = re.search(r'(\d+)일', duration_str)
                if day_part:
                    days += int(day_part.group(1))
            return days
        except:
            return 0
    
    def _create_detail_sheet(self, ws, report_data: Dict):
        """학생별 상세 명단 시트 (입소/퇴소 분리)"""
        detailed = report_data["detailed_list"]
        
        if not detailed:
            ws['A1'] = "데이터가 없습니다."
            return
        
        # 입소 학생과 퇴소 학생으로 분리
        enrolled_students = [s for s in detailed if s.get("재원상태") == "재원중"]
        discharged_students = [s for s in detailed if s.get("재원상태") == "퇴원"]
        
        # 퇴소 학생 재원기간 평균 계산
        discharged_durations = []
        for student in discharged_students:
            duration_str = student.get("재원기간", "")
            days = self._parse_duration_to_days(duration_str)
            if days > 0:
                discharged_durations.append(days)
        
        avg_duration_days = sum(discharged_durations) / len(discharged_durations) if discharged_durations else 0
        if avg_duration_days >= 30:
            avg_months = int(avg_duration_days // 30)
            avg_remaining_days = int(avg_duration_days % 30)
            if avg_remaining_days > 0:
                avg_duration_str = f"{avg_months}개월 {avg_remaining_days}일"
            else:
                avg_duration_str = f"{avg_months}개월"
        else:
            avg_duration_str = f"{int(avg_duration_days)}일"
        
        # 제목
        ws.merge_cells('A1:G1')
        title = ws['A1']
        title.value = "👥 학생별 상세 명단"
        title.font = Font(size=14, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 25
        
        current_row = 3
        
        # 입소 학생과 퇴소 학생 데이터 준비
        df_enrolled = pd.DataFrame(enrolled_students) if enrolled_students else pd.DataFrame()
        df_discharged = pd.DataFrame(discharged_students) if discharged_students else pd.DataFrame()
        
        # 입소 학생 데이터에서 퇴소일자와 퇴원사유 컬럼 제거
        if not df_enrolled.empty:
            columns_to_drop = ['퇴소일자', '퇴원사유']
            existing_columns_to_drop = [col for col in columns_to_drop if col in df_enrolled.columns]
            if existing_columns_to_drop:
                df_enrolled = df_enrolled.drop(columns=existing_columns_to_drop)
        
        # 퇴소 학생 컬럼 수
        discharged_cols = len(df_discharged.columns) if not df_discharged.empty else 0
        # 입소 학생 시작 열 (퇴소 학생 컬럼 + 간격 2열)
        enrolled_start_col = discharged_cols + 3 if discharged_cols > 0 else 1
        
        # ===== 퇴소 학생 섹션 (왼쪽) =====
        if not df_discharged.empty:
            # 퇴소 학생 섹션 제목
            ws.merge_cells(f'A{current_row}:{get_column_letter(discharged_cols)}{current_row}')
            section_title = ws[f'A{current_row}']
            section_title.value = "📌 퇴소 학생"
            section_title.font = Font(size=12, bold=True, color="FFFFFF")
            section_title.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            section_title.alignment = Alignment(horizontal="left")
            current_row += 1
            
            # 헤더
            for col_num, column in enumerate(df_discharged.columns, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = column
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # 재원기간 평균 표시 (G2 셀)
            avg_cell = ws['G2']
            avg_cell.value = f"평균 재원기간: {avg_duration_str}"
            avg_cell.font = Font(size=11, bold=True, color="FFFFFF")
            avg_cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
            avg_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            current_row += 1
            
            # 데이터
            discharged_data_start_row = current_row
            for row_idx, row_data in enumerate(df_discharged.values):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.value = value if value is not None else "-"
                    cell.alignment = Alignment(horizontal="left")
                    
                    # 재원상태 색상
                    if df_discharged.columns[col_num - 1] == "재원상태":
                        cell.font = Font(color="FF0000", bold=True)
                    
                    # 스트라이프
                    if current_row % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", 
                                               end_color="F2F2F2", fill_type="solid")
                current_row += 1
            
            discharged_data_end_row = current_row - 1
        
        # ===== 입소 학생 섹션 (오른쪽) =====
        if not df_enrolled.empty:
            # 입소 학생 섹션 제목
            enrolled_cols = len(df_enrolled.columns)
            ws.merge_cells(f'{get_column_letter(enrolled_start_col)}{3}:{get_column_letter(enrolled_start_col + enrolled_cols - 1)}{3}')
            section_title = ws[f'{get_column_letter(enrolled_start_col)}{3}']
            section_title.value = "📌 입소 학생 (재원중)"
            section_title.font = Font(size=12, bold=True, color="FFFFFF")
            section_title.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            section_title.alignment = Alignment(horizontal="left")
            
            # 헤더 행 설정
            header_row = 4
            
            # 헤더
            for col_num, column in enumerate(df_enrolled.columns, 1):
                col = enrolled_start_col + col_num - 1
                cell = ws.cell(row=header_row, column=col)
                cell.value = column
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # 데이터 (퇴소 학생과 같은 행에 맞춰서)
            data_start_row = header_row + 1
            if not df_discharged.empty:
                # 퇴소 학생 데이터 시작 행과 맞춤
                data_start_row = discharged_data_start_row
            
            for row_idx, row_data in enumerate(df_enrolled.values):
                data_row = data_start_row + row_idx
                for col_num, value in enumerate(row_data, 1):
                    col = enrolled_start_col + col_num - 1
                    cell = ws.cell(row=data_row, column=col)
                    cell.value = value if value is not None else "-"
                    cell.alignment = Alignment(horizontal="left")
                    
                    # 재원상태 색상
                    if df_enrolled.columns[col_num - 1] == "재원상태":
                        cell.font = Font(color="00B050", bold=True)
                    
                    # 스트라이프
                    if data_row % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", 
                                               end_color="F2F2F2", fill_type="solid")
            
            # 최종 행 업데이트
            if not df_discharged.empty:
                current_row = max(current_row, data_start_row + len(df_enrolled))
            else:
                current_row = data_start_row + len(df_enrolled)
        
        # 열 너비 자동 조정 (글자에 맞춰서)
        max_col = max(discharged_cols, enrolled_start_col + len(df_enrolled.columns) - 1) if not df_enrolled.empty else discharged_cols
        
        # 한글 문자를 고려한 너비 계산 함수
        def calculate_text_width(text):
            """한글과 영문을 고려한 텍스트 너비 계산"""
            if not text:
                return 0
            text_str = str(text)
            width = 0
            for char in text_str:
                # 한글, 한자 등 전각 문자는 2로 계산
                if ord(char) > 127:
                    width += 2
                else:
                    width += 1
            return width
        
        for col_num in range(1, max_col + 1):
            max_width = 0
            column_letter = get_column_letter(col_num)
            
            # J, K열은 최소 너비로 설정
            if column_letter in ['J', 'K']:
                ws.column_dimensions[column_letter].width = 3  # 최소 너비
                continue
            
            # P열은 평균 재원기간 셀(P2) 크기에 맞춰서 조정
            if column_letter == 'G':
                # P2 셀의 텍스트 크기 확인
                g2_cell = ws['G2']
                if g2_cell.value:
                    text_width = calculate_text_width(g2_cell.value)
                    ws.column_dimensions[column_letter].width = text_width + 2
                else:
                    ws.column_dimensions[column_letter].width = 10  # 기본값
                continue
            
            # S열은 글자 크기에 맞춰서 자동 조정
            if column_letter == 'S':
                for row in ws.iter_rows(min_row=3, max_row=current_row, min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            if cell.value:
                                text_width = calculate_text_width(cell.value)
                                if text_width > max_width:
                                    max_width = text_width
                        except:
                            pass
                # S열 너비 설정 (텍스트 너비 + 여유 공간 2)
                if max_width > 0:
                    ws.column_dimensions[column_letter].width = max_width + 2
                else:
                    ws.column_dimensions[column_letter].width = 10  # 기본값
                continue
            
            # 나머지 열은 일반 자동 조정
            for row in ws.iter_rows(min_row=3, max_row=current_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    try:
                        if cell.value:
                            text_width = calculate_text_width(cell.value)
                            if text_width > max_width:
                                max_width = text_width
                    except:
                        pass
            # 열 너비 설정 (텍스트 너비 + 여유 공간 2)
            if max_width > 0:
                ws.column_dimensions[column_letter].width = max_width + 2
            else:
                ws.column_dimensions[column_letter].width = 10  # 기본값


####

    
     
      

####

class ReportOrchestrator:
    def __init__(self):
        self.notion = NotionManager()
        self.ai = OllamaAnalyzer()
        self.discharge_report = EnhancedDischargeReportGenerator(self.notion)
        #self.pdf = PDFConverter()
        #self.security = SecurityManager()
        #self.file_manager = LocalFileManager()
    
    async def _process_discharge_report(self, query_results: Dict, query: ReportQuery):
        """입퇴소 보고서 (차트 포함)"""
        
        if isinstance(query, list):
            q0 = query[0]
        else:
            q0 = query

        teacher_name = q0.filters.get("teacher_name")
        
        # 날짜 범위에서 년월 추출
        year = datetime.now().year
        month = datetime.now().month
        
        if q0.date_range:
            try:
                # date_range의 end 날짜를 기준으로 년월 추출 (또는 start 날짜)
                date_str = q0.date_range.get("end") or q0.date_range.get("start")
                if date_str:
                    if "T" in date_str:
                        date_obj = datetime.fromisoformat(date_str.split("T")[0])
                    else:
                        date_obj = datetime.fromisoformat(date_str)
                    year = date_obj.year
                    month = date_obj.month
            except Exception as e:
                logger.warning(f"⚠️ 날짜 범위 파싱 실패, 현재 날짜 사용: {str(e)}")

        # 보고서 데이터 생성
        report_data = await self.discharge_report.generate_monthly_report(
            query_results, teacher_name, year, month
        )
        
        # 차트 포함 Excel 생성
        filename = f"discharge_chart_{teacher_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        excel_path = self.discharge_report.create_excel_with_chart(
            report_data, filename
        )

    async def process_request(self, request: ReportRequest):
        """보고서 요청 전체 처리"""
        logger.info(f"\n{'='*60}")
        logger.info(f"🔨 처리 시작: {request.requester_name}님의 요청")
        logger.info(f"   질문: {request.question}")
        logger.info(f"{'='*60}\n")
        
        try:
            # 상태 업데이트: 검토중
            await self.notion.update_request_status(request.id, "검토중")

            # 1. 자연어 질문 분석 -> 쿼리 생성
            query = await self.ai.analyze_question(request.question)
            
            # 2. 쿼리 실행 및 데이터 수집
            query_results = await self.notion.query_multiple_tables(query)
            
            # 3. 보고서 생성 및 전달
            await self._process_discharge_report(query_results, query)

            # 완료 상태로 업데이트
            await self.notion.update_request_status(request.id, "완료됨")
            logger.info(f"\n✅ 처리 완료!\n{'='*60}\n")
            return
        except Exception as e:
            logger.error(f"❌ 처리 실패: {str(e)}")
            await self.notion.update_request_status(
                request.id, "실패", error=str(e)
            )

####

@dataclass
class DataImportRequest:
    """데이터 입력 요청"""
    table_type: str
    dataframe: pd.DataFrame
    file_infos: List[Dict[str, Any]] = field(default_factory=list)  # 처리할 파일 정보 목록
    id: str = field(default_factory=lambda: f"data_import_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}")
    _is_data_import: bool = True
    _retry_count: int = 0

class PollingSystem:
    def __init__(self):
        self.orchestrator = ReportOrchestrator()
        self.is_running = False
        self.queue = asyncio.PriorityQueue()  # 우선순위 큐로 변경
        self.processed_ids = set()  # 처리 중이거나 완료된 요청 ID 추적
        self.worker_tasks = []  # 여러 워커 태스크 저장
        self.polling_task = None
        self.is_processing_report = False  # 보고서 생성 중 플래그
        self.report_lock = asyncio.Lock()  # 보고서 생성 락
        self._queue_order = 0  # 큐에 추가된 순서 (우선순위가 같을 때 비교용)
    
    async def _worker(self):
        """큐에서 요청을 하나씩 꺼내서 처리하는 워커 (우선순위 기반)"""
        logger.info("👷 워커 시작")
        while self.is_running:
            request = None
            priority = None
            try:
                # 큐에서 요청 가져오기 (타임아웃 1초, 우선순위 큐)
                try:
                    priority, order, request = await asyncio.wait_for(self.queue.get(), timeout=1.0)
                except asyncio.TimeoutError:
                    continue
                
                if request is None:
                    continue
                
                # 이미 처리 중이거나 완료된 요청은 건너뛰기
                if request.id in self.processed_ids:
                    logger.debug(f"⏭️ 이미 처리된 요청 건너뛰기: {request.id}")
                    self.queue.task_done()
                    continue
                
                # 데이터 입력 작업인 경우 보고서 생성이 끝날 때까지 대기
                is_data_import = getattr(request, '_is_data_import', False)
                if is_data_import and self.is_processing_report:
                    logger.info(f"⏳ 데이터 입력 대기 중: 보고서 생성 완료 대기... (요청 ID: {request.id})")
                    # 보고서 생성이 끝날 때까지 대기
                    while self.is_processing_report and self.is_running:
                        await asyncio.sleep(0.5)
                    logger.info(f"✅ 데이터 입력 시작: 보고서 생성 완료됨 (요청 ID: {request.id})")
                
                logger.info(f"📝 큐에서 요청 가져옴: {request.id} (우선순위: {priority}, 큐 크기: {self.queue.qsize()})")
                
                # 처리 시도
                try:
                    # 처리 시작 시 processed_ids에 추가 (중복 처리 방지)
                    self.processed_ids.add(request.id)
                    
                    # 보고서 생성 작업인 경우 플래그 설정
                    if not is_data_import:
                        async with self.report_lock:
                            self.is_processing_report = True
                            logger.info(f"📊 보고서 생성 시작: {request.id}")
                    
                    # 데이터 입력 작업 처리
                    if is_data_import:
                        await self._process_data_import(request)
                    else:
                        # 보고서 생성 작업 처리
                        await self.orchestrator.process_request(request)
                    
                    logger.info(f"✅ 요청 처리 완료: {request.id}")
                except Exception as e:
                    logger.error(f"❌ 요청 처리 실패: {request.id}, 오류: {str(e)}")
                    # 처리 실패 시 processed_ids에서 제거하여 재시도 가능하도록
                    self.processed_ids.discard(request.id)
                    import traceback
                    logger.error(f"상세 오류:\n{traceback.format_exc()}")
                    # 실패한 요청을 다시 큐에 넣어 재시도 (무한 루프 방지를 위해 최대 3회)
                    retry_count = getattr(request, '_retry_count', 0)
                    if retry_count < 3:
                        request._retry_count = retry_count + 1
                        # 재시도 시에도 순서 번호 증가
                        self._queue_order += 1
                        await self.queue.put((priority, self._queue_order, request))  # 우선순위 유지
                        logger.info(f"🔄 요청 재시도 큐에 추가: {request.id} (재시도 {retry_count + 1}/3)")
                finally:
                    # 보고서 생성 작업인 경우 플래그 해제
                    if not is_data_import:
                        async with self.report_lock:
                            self.is_processing_report = False
                            logger.info(f"📊 보고서 생성 완료: {request.id}")
                    # 큐 작업 완료 표시 (성공/실패 관계없이)
                    self.queue.task_done()
                    
            except Exception as e:
                logger.error(f"❌ 워커 에러: {str(e)}")
                import traceback
                logger.error(f"상세 오류:\n{traceback.format_exc()}")
                if request:
                    self.queue.task_done()
                await asyncio.sleep(1)
    
    async def _process_data_import(self, request: DataImportRequest):
        """데이터 입력 작업 처리"""
        logger.info(f"📤 [{request.table_type}] Notion DB에 데이터 추가 시작: {len(request.dataframe)}개 행")
        
        added_count = 0
        failed_count = 0
        
        for idx, row in request.dataframe.iterrows():
            try:
                # DataFrame 행을 Notion 속성으로 변환
                properties = excel_importer._convert_dataframe_row_to_notion_properties(
                    row, request.table_type, request.dataframe
                )
                
                if not properties:
                    logger.warning(f"⚠️ [{request.table_type}] 행 {idx}: 변환된 속성이 없어 건너뜁니다.")
                    continue
                
                # Notion에 추가
                await self.orchestrator.notion.client.pages.create(
                    parent={"database_id": self.orchestrator.notion.db_map[request.table_type]},
                    properties=properties
                )
                
                added_count += 1
                if added_count % 10 == 0:
                    logger.info(f"📝 [{request.table_type}] 진행 중: {added_count}/{len(request.dataframe)}개 추가됨")
                
                # API 제한 고려 (초당 3회)
                await asyncio.sleep(0.35)
                
            except Exception as e:
                failed_count += 1
                logger.error(f"❌ [{request.table_type}] 행 {idx} 추가 실패: {e}")
        
        logger.info(f"✅ [{request.table_type}] Notion DB 추가 완료: 성공 {added_count}개, 실패 {failed_count}개")
        
        # 처리된 파일을 imported 폴더로 이동 (해당 요청에 포함된 파일들만 이동)
        if request.file_infos:
            moved_count = excel_handler.move_specific_files_to_imported(request.table_type, request.file_infos)
            logger.info(f"✅ [{request.table_type}] {moved_count}개 파일을 imported 폴더로 이동 완료")
        else:
            logger.warning(f"⚠️ [{request.table_type}] 이동할 파일 정보가 없습니다.")
    
    async def _polling(self, interval: int = 30):
        """주기적으로 새로운 요청을 큐에 추가하는 폴링 태스크"""
        logger.info("🔍 폴링 시작")
        
        # 초기화: 대기중인 모든 요청을 큐에 추가
        try:
            initial_requests = await self.orchestrator.notion.get_pending_requests()
            for req in initial_requests:
                # 이미 처리된 요청이지만 상태가 다시 "대기중"으로 변경된 경우 재처리
                if req.id in self.processed_ids:
                    logger.info(f"🔄 초기화: 재처리 요청 발견 (상태가 다시 대기중으로 변경됨): {req.id}")
                    self.processed_ids.discard(req.id)  # processed_ids에서 제거하여 재처리 가능하도록
                
                if req.id not in self.processed_ids:
                    # 보고서 생성 요청은 우선순위 2 (낮음)
                    self._queue_order += 1
                    await self.queue.put((2, self._queue_order, req))
                    logger.info(f"📥 초기 요청 큐에 추가: {req.id} (우선순위: 2, 큐 크기: {self.queue.qsize()})")
            logger.info(f"✅ 초기 {len(initial_requests)}개 요청 큐에 추가 완료")
        except Exception as e:
            logger.error(f"❌ 초기 요청 로드 실패: {str(e)}")
        
        # 주기적으로 새로운 요청 확인
        while self.is_running:
            try:
                await asyncio.sleep(interval)
                
                if not self.is_running:
                    break
                
                requests = await self.orchestrator.notion.get_pending_requests()
                
                # 새로운 요청만 큐에 추가 (processed_ids에 추가하지 않음 - 워커에서 처리할 때 추가)
                new_count = 0
                for req in requests:
                    # 이미 처리된 요청이지만 상태가 다시 "대기중"으로 변경된 경우 재처리
                    if req.id in self.processed_ids:
                        logger.info(f"🔄 재처리 요청 발견 (상태가 다시 대기중으로 변경됨): {req.id}")
                        self.processed_ids.discard(req.id)  # processed_ids에서 제거하여 재처리 가능하도록
                    
                    if req.id not in self.processed_ids:
                        # 보고서 생성 요청은 우선순위 2 (낮음)
                        self._queue_order += 1
                        await self.queue.put((2, self._queue_order, req))
                        new_count += 1
                        logger.info(f"📥 새 요청 큐에 추가: {req.id} (우선순위: 2, 큐 크기: {self.queue.qsize()})")
                
                if new_count == 0:
                    logger.info(f"💤 새 요청 없음 (큐 크기: {self.queue.qsize()}) ({datetime.now().strftime('%H:%M:%S')})")
                else:
                    logger.info(f"📥 {new_count}개 새 요청 큐에 추가됨 (큐 크기: {self.queue.qsize()})")
                
            except Exception as e:
                logger.error(f"❌ 폴링 에러: {str(e)}")
    
    async def start(self, interval: int = 30, num_workers: int = 1):
        self.is_running = True
        logger.info("🚀 학원 보고서 시스템 시작")
        logger.info(f"⏰ 폴링 간격: {interval}초")
        logger.info(f"👷 워커 수: {num_workers}개")
        logger.info("-" * 60)
        
        # 워커 태스크 시작 (큐에서 요청 처리)
        self.worker_tasks = []
        for i in range(num_workers):
            task = asyncio.create_task(self._worker())
            self.worker_tasks.append(task)
            logger.info(f"👷 워커 {i+1} 시작")
        
        # 폴링 태스크 시작 (새 요청을 큐에 추가)
        self.polling_task = asyncio.create_task(self._polling(interval))
        
        # 모든 태스크가 완료될 때까지 대기
        try:
            await asyncio.gather(*self.worker_tasks, self.polling_task)
        except asyncio.CancelledError:
            pass
    
    def stop(self):
        self.is_running = False
        logger.info("⏹️ 시스템 중지 중...")
        
        # 모든 워커 태스크 취소
        for task in self.worker_tasks:
            if task:
                task.cancel()
        if self.polling_task:
            self.polling_task.cancel()
        
        logger.info(f"⏹️ 시스템 중지 완료 (큐에 남은 요청: {self.queue.qsize()}개)")


####

app = FastAPI(title="학원 보고서 시스템")
polling = PollingSystem()

# ExcelFileHandler 인스턴스 생성 (NotionManager 주입)
notion_manager = NotionManager()
excel_handler = ExcelFileHandler(notion_manager=notion_manager)
excel_importer = ExcelImporter(notion_manager)

@app.on_event("startup")
async def startup():
    """서버 시작 시 폴링 및 엑셀 파일 감시 시작"""
    asyncio.create_task(polling.start(interval=30))
    asyncio.create_task(excel_file_watcher_worker())

async def excel_file_watcher_worker():
    """엑셀 파일 감시 및 자동 처리 백그라운드 워커
    
    10초마다 다음 작업을 자동으로 수행:
    1. 새 엑셀 파일 감지 및 저장
    2. 전처리 (합치기, 중복 제거, 필터링)
    3. Notion DB에 추가
    4. 처리된 파일을 imported 폴더로 이동
    """
    logger.info("📂 엑셀 파일 자동 처리 워커 시작")
    
    while True:
        try:
            # 1. 새 엑셀 파일 감지 및 저장
            result = excel_handler.watch_and_store()
            
            # 각 폴더별로 처리
            for table_type in ["class", "discharge", "student"]:
                if table_type in result and len(result[table_type]) > 0:
                    try:
                        logger.info(f"🔄 [{table_type}] 자동 처리 시작...")
                        
                        # 2. 전처리
                        df = await excel_handler.preprocess_and_merge(table_type)
                        
                        if df is None or df.empty:
                            logger.info(f"⚠️ [{table_type}] 전처리 후 데이터가 없지만 파일은 이동합니다.")
                            # 데이터가 없어도 파일을 imported 폴더로 이동
                            moved_count = excel_handler.move_processed_files_to_imported(table_type)
                            logger.info(f"✅ [{table_type}] {moved_count}개 파일을 imported 폴더로 이동 완료")
                            continue
                        
                        logger.info(f"✅ [{table_type}] 전처리 완료: {len(df)}개 행")
                        
                        # 3. 큐에 추가하기 전에 해당 테이블 타입의 파일들을 queued_files에 추가
                        for file_info in excel_handler.stored_files[table_type]:
                            file_key = file_info.get("file_key")
                            if file_key:
                                excel_handler.queued_files.add(file_key)
                        
                        # 4. 데이터 입력 작업을 우선순위 큐에 추가 (우선순위 1 = 높음)
                        # 현재 stored_files에 있는 파일 정보를 복사 (처리 완료 후 이동하기 위해)
                        current_file_infos = excel_handler.stored_files[table_type].copy()
                        data_import_request = DataImportRequest(
                            table_type=table_type,
                            dataframe=df,
                            file_infos=current_file_infos  # 파일 정보 포함
                        )
                        polling._queue_order += 1
                        await polling.queue.put((1, polling._queue_order, data_import_request))  # 우선순위 1
                        logger.info(f"📥 [{table_type}] 데이터 입력 작업 큐에 추가됨 (우선순위: 1, 큐 크기: {polling.queue.qsize()})")
                        
                        # 큐에 추가한 후 해당 파일들을 stored_files에서 제거 (처리 완료 후 이동하기 위해 요청에 포함됨)
                        # 요청에 파일 정보가 포함되어 있으므로, 처리 완료 시 해당 파일들만 이동됨
                        for file_info in current_file_infos:
                            file_key = file_info.get("file_key")
                            if file_key:
                                # stored_files에서 해당 파일 제거
                                excel_handler.stored_files[table_type] = [
                                    f for f in excel_handler.stored_files[table_type]
                                    if f.get("file_key") != file_key
                                ]
                        logger.debug(f"🔄 [{table_type}] 큐 추가 후 해당 파일들을 stored_files에서 제거 (요청에 포함됨)")
                            
                    except Exception as e:
                        logger.error(f"❌ [{table_type}] 자동 처리 오류: {str(e)}")
                        import traceback
                        logger.error(traceback.format_exc())
                        
        except Exception as e:
            logger.error(f"❌ 엑셀 파일 감시 오류: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
        
        # 10초마다 스캔
        await asyncio.sleep(10)

@app.get("/")
async def root():
    return {
        "service": "학원 보고서 자동 생성 시스템",
        "status": "running",
        "ai": "Ollama (qwen3:8b)"
    }

@app.get("/health")
async def health():
    return {"status": "healthy"}

@app.get("/download/{date}/{filename}")
async def download_file(date: str, filename: str):
    """파일 다운로드 엔드포인트"""
    from fastapi.responses import FileResponse
    
    file_path = config.REPORTS_DIR / date / filename
    
    if not file_path.exists():
        return {"error": "File not found"}, 404
    
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type='application/octet-stream'
    )

@app.post("/webhook/notion")
async def webhook():
    """Notion 웹훅 (실시간 처리용) - 큐에 추가"""
    requests = await polling.orchestrator.notion.get_pending_requests()
    added_count = 0
    for req in requests:
        # 이미 처리된 요청이지만 상태가 다시 "대기중"으로 변경된 경우 재처리
        if req.id in polling.processed_ids:
            logger.info(f"🔄 웹훅: 재처리 요청 발견 (상태가 다시 대기중으로 변경됨): {req.id}")
            polling.processed_ids.discard(req.id)  # processed_ids에서 제거하여 재처리 가능하도록
        
        if req.id not in polling.processed_ids:
            # 보고서 생성 요청은 우선순위 2 (낮음)
            polling._queue_order += 1
            await polling.queue.put((2, polling._queue_order, req))
            # processed_ids에 추가하지 않음 - 워커에서 처리할 때 추가
            added_count += 1
            logger.info(f"📥 웹훅으로 새 요청 큐에 추가: {req.id} (우선순위: 2, 큐 크기: {polling.queue.qsize()})")
    return {"status": "processing", "added_to_queue": added_count}

@app.get("/excel/watch")
async def watch_excel_files():
    """input 폴더의 새 엑셀 파일을 감지하고 폴더별로 구별해서 저장"""
    try:
        result = excel_handler.watch_and_store()
        
        # 결과 요약 (폴더별)
        summary = {}
        data_by_folder = {}
        
        for folder_name, file_list in result.items():
            summary[folder_name] = {
                "file_count": len(file_list),
                "total_rows": sum(file_info["rows"] for file_info in file_list)
            }
            
            # 폴더별 파일 정보 (DataFrame 제외)
            data_by_folder[folder_name] = [
                {
                    "file_name": file_info["file_name"],
                    "file_path": file_info["file_path"],
                    "folder": file_info["folder"],
                    "rows": file_info["rows"],
                    "columns": file_info["columns"],
                    "read_time": file_info["read_time"]
                }
                for file_info in file_list
            ]
        
        return {
            "status": "success",
            "summary": summary,
            "data_by_folder": data_by_folder
        }
    except Exception as e:
        logger.error(f"❌ 엑셀 파일 감시 오류: {str(e)}")
        return {"status": "error", "message": str(e)}

@app.get("/excel/stored")
async def get_stored_files(table_type: Optional[str] = None):
    """저장된 엑셀 파일 정보 조회 (폴더별로 구별)"""
    try:
        result = excel_handler.get_stored_files(table_type)
        
        # 결과 요약
        summary = {}
        data_by_folder = {}
        
        for folder_name, file_list in result.items():
            summary[folder_name] = {
                "file_count": len(file_list),
                "total_rows": sum(file_info["rows"] for file_info in file_list)
            }
            
            # 폴더별 파일 정보 (DataFrame 제외)
            data_by_folder[folder_name] = [
                {
                    "file_name": file_info["file_name"],
                    "file_path": file_info["file_path"],
                    "folder": file_info["folder"],
                    "rows": file_info["rows"],
                    "columns": file_info["columns"],
                    "read_time": file_info["read_time"]
                }
                for file_info in file_list
            ]
        
        return {
            "status": "success",
            "summary": summary,
            "data_by_folder": data_by_folder
        }
    except Exception as e:
        logger.error(f"❌ 저장된 파일 조회 오류: {str(e)}")
        return {"status": "error", "message": str(e)}

@app.post("/excel/clear")
async def clear_stored_files(table_type: Optional[str] = None):
    """저장된 파일 정보 초기화"""
    excel_handler.clear_stored_files(table_type)
    return {
        "status": "success",
        "message": f"{table_type if table_type else '모든'} 폴더의 저장된 파일 정보가 초기화되었습니다."
    }

@app.post("/excel/reset")
async def reset_excel_handler():
    """처리된 파일 목록 초기화 (모든 파일을 다시 읽을 수 있도록)"""
    excel_handler.reset_processed_files()
    return {"status": "success", "message": "처리된 파일 목록이 초기화되었습니다."}

@app.post("/excel/preprocess/{table_type}")
async def preprocess_excel_files(table_type: str):
    """저장된 엑셀 파일들을 합치고 중복 제거 및 날짜 필터링 (전처리)
    
    Args:
        table_type: 테이블 타입 (class, discharge, student)
    """
    try:
        df = await excel_handler.preprocess_and_merge(table_type)
        
        if df is None:
            return {
                "status": "error",
                "message": f"{table_type} 폴더에 처리할 파일이 없습니다."
            }
        
        return {
            "status": "success",
            "table_type": table_type,
            "rows": len(df),
            "columns": list(df.columns),
            "message": f"{table_type} 폴더의 {len(excel_handler.stored_files[table_type])}개 파일을 합쳐서 {len(df)}개 행으로 전처리 완료"
        }
    except Exception as e:
        logger.error(f"❌ 전처리 오류: {str(e)}")
        return {"status": "error", "message": str(e)}

@app.post("/excel/import/{table_type}")
async def import_excel_to_notion(table_type: str):
    """전처리된 엑셀 파일을 Notion에 추가하고 처리된 파일을 imported 폴더로 이동
    
    Args:
        table_type: 테이블 타입 (class, discharge, student)
    """
    try:
        # 1. 전처리
        logger.info(f"🔄 [{table_type}] 전처리 시작...")
        df = await excel_handler.preprocess_and_merge(table_type)
        
        if df is None or df.empty:
            return {
                "status": "error",
                "message": f"{table_type} 폴더에 처리할 파일이 없거나 전처리 후 데이터가 없습니다."
            }
        
        logger.info(f"✅ [{table_type}] 전처리 완료: {len(df)}개 행")
        
        # 2. Notion에 추가
        logger.info(f"📤 [{table_type}] Notion DB에 추가 시작...")
        added_count = await excel_importer.add_preprocessed_data_to_notion(df, table_type)
        
        if added_count == 0:
            return {
                "status": "error",
                "message": f"{table_type} 데이터를 Notion에 추가하지 못했습니다."
            }
        
        logger.info(f"✅ [{table_type}] Notion DB에 {added_count}개 추가 완료")
        
        # 3. 처리된 파일을 imported 폴더로 이동
        logger.info(f"📦 [{table_type}] 처리된 파일을 imported 폴더로 이동 시작...")
        moved_count = excel_handler.move_processed_files_to_imported(table_type)
        
        logger.info(f"✅ [{table_type}] {moved_count}개 파일을 imported 폴더로 이동 완료")
        
        return {
            "status": "success",
            "table_type": table_type,
            "preprocessed_rows": len(df),
            "notion_added": added_count,
            "files_moved": moved_count,
            "message": f"{table_type} 처리 완료: {added_count}개 데이터 Notion 추가, {moved_count}개 파일 이동"
        }
    except Exception as e:
        logger.error(f"❌ [{table_type}] import 오류: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return {"status": "error", "message": str(e)}

@app.post("/excel/import-all")
async def import_all_excel_to_notion():
    """모든 폴더의 전처리된 엑셀 파일을 Notion에 추가하고 처리된 파일을 imported 폴더로 이동"""
    try:
        result = {}
        
        for table_type in ["class", "discharge", "student"]:
            try:
                # 1. 전처리
                logger.info(f"🔄 [{table_type}] 전처리 시작...")
                df = await excel_handler.preprocess_and_merge(table_type)
                
                if df is None or df.empty:
                    result[table_type] = {
                        "status": "skipped",
                        "message": "처리할 파일이 없거나 전처리 후 데이터가 없습니다."
                    }
                    continue
                
                logger.info(f"✅ [{table_type}] 전처리 완료: {len(df)}개 행")
                
                # 2. Notion에 추가
                logger.info(f"📤 [{table_type}] Notion DB에 추가 시작...")
                added_count = await excel_importer.add_preprocessed_data_to_notion(df, table_type)
                
                # 3. 처리된 파일을 imported 폴더로 이동
                logger.info(f"📦 [{table_type}] 처리된 파일을 imported 폴더로 이동 시작...")
                moved_count = excel_handler.move_processed_files_to_imported(table_type)
                
                result[table_type] = {
                    "status": "success",
                    "preprocessed_rows": len(df),
                    "notion_added": added_count,
                    "files_moved": moved_count
                }
                
            except Exception as e:
                logger.error(f"❌ [{table_type}] 처리 오류: {str(e)}")
                result[table_type] = {
                    "status": "error",
                    "message": str(e)
                }
        
        return {
            "status": "success",
            "results": result
        }
    except Exception as e:
        logger.error(f"❌ 전체 import 오류: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return {"status": "error", "message": str(e)}

@app.post("/excel/preprocess-all")
async def preprocess_all_excel_files():
    """모든 폴더의 저장된 엑셀 파일들을 전처리 (합치기 + 중복 제거 + 날짜 필터링)"""
    try:
        result = await excel_handler.preprocess_all_folders()
        
        summary = {}
        for table_type, df in result.items():
            if df is not None:
                summary[table_type] = {
                    "rows": len(df),
                    "columns": list(df.columns),
                    "file_count": len(excel_handler.stored_files[table_type])
                }
            else:
                summary[table_type] = {
                    "rows": 0,
                    "message": "처리할 파일이 없습니다."
                }
        
        return {
            "status": "success",
            "summary": summary
        }
    except Exception as e:
        logger.error(f"❌ 전처리 오류: {str(e)}")
        return {"status": "error", "message": str(e)}


####

if __name__ == "__main__":
    uvicorn.run(
        app,
        host="0.0.0.0",  # 외부 접근 허용
        port=8000,
        log_level="info"
    )
